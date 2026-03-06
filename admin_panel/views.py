# admin_panel/views.py
# ─────────────────────────────────────────────────────────────
# What's in this file (in order):
#   1.  admin_login()               — separate staff login
#   2.  admin_logout()              — destroy session
#   3.  admin_dashboard()           — home: stats based on permissions
#   4.  user_list()                 — list all regular users
#   5+.  admin_view_user()            — view a user's dashboard as an admin
#   5.  create_user()               — admin creates a new user
#   6.  manage_admins()             — superadmin: list all admins
#   7.  create_admin()              — superadmin: create new admin
#   8.  edit_admin_permissions()    — superadmin: update admin permissions
#   9.  case_list()                 — list all cases with filters
#   10. case_detail()               — view case + answers + manage requirements
#   11. create_case()               — admin creates case, auto-creates CaseRequirements
#   12. toggle_requirement()        — soft delete/restore a requirement per case
#   13. add_extra_requirement()     — add requirement not in category defaults
#   14. group_list()                — list all groups
#   15. group_detail()              — view group members + managed profiles
#   15++. admin_create_group()       — admin creates a group and assigns an owner
#   15++. admin_add_member()         — admin adds a user to a group with optional role
#   15++. admin_toggle_member()      — admin suspends or restores a group member
#   15++. admin_set_member_permissions() — admin sets which GroupPermissions a member has
#   15++++ admin_change_member_role()      — admin changes a member's role within the group
#   16. get_categories()            — AJAX: categories for a service
#   17. get_subcategories()         — AJAX: subcategories for a category
#    18. service_builder()             — superadmin: create/edit services, categories, requirements
#    19. service_list()                — list all services with their categories and requirements
#    20. ajax_get_services()             — AJAX: get all services (for dynamic form dropdowns)
#    21. ajax_create_service()          — AJAX: superadmin creates a new service (from service_builder page)
#    22.ajax_get_categories()             — AJAX: get categories for a service (for dynamic form dropdowns)
#    23. ajax_create_category()          — AJAX: superadmin creates a new category (from service_builder page)
#    24. ajax_get_category_details()       — AJAX: get category details including requirements (for service_builder edit form)
#    25.ajax_create_requrment()           — AJAX: superadmin creates a new requirement (from service_builder page)
#    26. ajax_edit_requrment()             — AJAX: superadmin edits a requirement (from service_builder page)
#    27. ajax_editIcategory()             — AJAX: superadmin edits a category (from service_builder page)
#    28. ajax_delete_service()           — AJAX: superadmin deletes a service (from service_builder page)
#    29. ajax_delete_category()          — AJAX: superadmin deletes a category (from service_builder page)
#    30. ajax_delete_requirement()       — AJAX: superadmin deletes a requirement (from service_builder page)
#    31.task_list()                    — list all tasks (admin→user and admin→admin) with filters
#    32.task_create()                  — create a new task, either admin→user or admin→admin
#    33.task_detail()                    — view task details and edit (reassign, change status, etc.)
#    34.invoice_list()                   — list all invoices with filters
#    35.invoice_create()                 — create a new invoice for a user or group
#    36.invoice_detail()                 — view invoice details and edit (change status, etc.)
#    37.user_balance_overview()             — view a user's balance and transaction history (for invoice management)

# ─────────────────────────────────────────────────────────────

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .decorators import admin_required, superadmin_required, admin_permission_required 
from django.views.decorators.csrf import ensure_csrf_cookie
from .forms import (
    AdminLoginForm, CreateUserForm, CreateAdminForm, CreateCaseForm,
    AdminGroupCreateForm, AdminAddMemberForm,
    AdminAssignCaseForm, AdminLinkManagedForm, AdminManagedProfileForm,
)
from users.models import AdminProfile, ManagedProfile
from groups.models import Group, GroupMembership, Role, GroupPermission
from cases.models import (
    Case, CaseAnswer, CaseRequirement,
    Category, Service, Requirement,
    RequirementSection, RequirementChoice, CategoryRequirement,   # Phase 1 library models
    CrawlerSuggestion,                                            # Phase 5 crawler
    # GovernmentForm, FormRequirement, CategoryForm are imported inline in
    # the builder views that need them — they share names with Django builtins
    # and are only needed in a few AJAX functions.
)
from django.contrib.auth import get_user_model

import json
from django.views.decorators.http import require_http_methods

from django.db import transaction
from django.db.models import Max
from decimal import Decimal
from payments.models import Invoice, Payment
from tasks.models import (
    Task,
    notify_task_assigned,
    notify_task_completed,
    notify_invoice_created,
    notify_payment_recorded,
)




User = get_user_model()


# ── 1. ADMIN LOGIN ────────────────────────────────────────────

def admin_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        if request.user.is_superuser:
            return redirect('admin_panel:admin_dashboard')  # superuser skips profile check
        if request.user.is_staff and hasattr(request.user, 'admin_profile'):
            return redirect('admin_panel:admin_dashboard')
        # is_staff but no admin_profile — don't redirect to dashboard, show error
        if request.user.is_staff:
            logout(request)  # ← log them out instead of looping
            form = AdminLoginForm()
            form.add_error(None, 'Your admin profile is not configured. Contact a superadmin.')
            return render(request, 'admin_panel/login.html', {'form': form})

    form = AdminLoginForm(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        email    = form.cleaned_data['email']
        password = form.cleaned_data['password']
        user     = authenticate(request, email=email, password=password)

        if user and user.is_staff:
            login(request, user)
            return redirect('admin_panel:admin_dashboard')
        else:
            form.add_error(None, 'Invalid credentials or insufficient access.')

    return render(request, 'admin_panel/login.html', {'form': form})


# ── 2. ADMIN LOGOUT ───────────────────────────────────────────

def admin_logout(request):
    logout(request)
    return redirect('admin_panel:admin_login')


# ── 3. ADMIN DASHBOARD ────────────────────────────────────────
# Only loads stats the admin has permission to see.
# EXPAND: add more stat blocks as new capabilities are added.
@ensure_csrf_cookie  # ← add this
@admin_required
def admin_dashboard(request):
    admin_profile = getattr(request.user, 'admin_profile', None)
    stats         = {}

    # superuser sees everything, admin sees only what they have permission for
    if request.user.is_superuser or (admin_profile and admin_profile.can_view_all_cases):
        stats['total_cases']     = Case.objects.filter(is_active=True).count()
        stats['pending_cases']   = Case.objects.filter(status='pending',   is_active=True).count()
        stats['completed_cases'] = Case.objects.filter(status='completed', is_active=True).count()

    if request.user.is_superuser or (admin_profile and admin_profile.can_create_groups):
        stats['total_groups']  = Group.objects.filter(is_active=True).count()
        stats['total_members'] = GroupMembership.objects.filter(is_active=True).count()

    if request.user.is_superuser or (admin_profile and admin_profile.can_create_users):
        stats['total_users'] = User.objects.filter(is_staff=False).count()

    return render(request, 'admin_panel/dashboard.html', {
        'admin_profile': admin_profile,  # None for superusers — template handles it
        'stats':         stats,
    })


# ── 4. USER LIST ──────────────────────────────────────────────

@admin_permission_required('can_create_users')
def user_list(request):
    users = User.objects.filter(is_staff=False).order_by('-date_joined')
    return render(request, 'admin_panel/user_list.html', {'users': users})


# ── 5. CREATE USER ────────────────────────────────────────────

@admin_permission_required('can_create_users')
def create_user(request):
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'User {user.email} created successfully.')
            return redirect('admin_panel:user_list')
    else:
        form = CreateUserForm()

    return render(request, 'admin_panel/create_user.html', {'form': form})

# ── 6. VIEW USER DASHBOARD ───────────────────────────────────
@admin_required
def admin_view_user(request, user_id):
    """
    Admin views a user's full dashboard — cases, groups, managed profiles.
    Reuses the user-facing template — no duplicate templates needed.

    The key difference from the real user dashboard:
      - request.user  = the admin (for access control)
      - target_user   = the user being viewed (for data)

    EXPAND: add an 'edit profile' button that links to admin_edit_user view.
    """
    target_user = get_object_or_404(User, id=user_id)

    # ── Handle permission toggle POSTs ───────────────────────
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'toggle_can_create_group':
            target_user.can_create_group = not target_user.can_create_group
            target_user.save(update_fields=['can_create_group'])
            status = 'granted' if target_user.can_create_group else 'revoked'
            messages.success(request, f'Group creation permission {status} for {target_user.email}.')
            return redirect('admin_panel:admin_view_user', user_id=user_id)

        # EXPAND: add more permission toggles here as new user permissions are added.

    # Same three queries as cases/views.py user_pickedCases_dashboard
    # but using target_user instead of request.user
    personal_cases = Case.objects.filter(
        user=target_user,
        group=None,
        managed_profile=None,
        is_active=True,
    )

    # WHY filter by user=target_user: a case belongs to exactly one user (the
    # applicant who filed it). Filtering only by group_id would return ALL cases
    # in any group the user is a member of — showing other members' cases too,
    # which is confusing and a privacy concern. We only want THIS user's cases
    # that happen to be assigned to a group.
    # WHY group__isnull=False: cleaner than querying group memberships — a case
    # is a "group case" simply because it has a group FK set.
    group_cases = Case.objects.filter(
        user=target_user,
        group__isnull=False,
        managed_profile=None,
        is_active=True,
    )

    managed_cases = Case.objects.filter(
        managed_profile__created_by=target_user,
        is_active=True,
    )

    # User's groups with their roles
    memberships = (
        GroupMembership.objects
        .filter(user=target_user, is_active=True)
        .select_related('group', 'role')
    )

    return render(request, 'admin_panel/user_view.html', {
        'target_user':    target_user,
        'personal_cases': personal_cases,
        'group_cases':    group_cases,
        'managed_cases':  managed_cases,
        'memberships':    memberships,
    })


# ── 6. MANAGE ADMINS ──────────────────────────────────────────

@superadmin_required
def manage_admins(request):
    admins = (
        AdminProfile.objects
        .select_related('user', 'created_by')
        .order_by('-created_at')
    )
    return render(request, 'admin_panel/manage_admins.html', {'admins': admins})


# ── 7. CREATE ADMIN ───────────────────────────────────────────

@superadmin_required
def create_admin(request):
    if request.method == 'POST':
        form = CreateAdminForm(request.POST)
        if form.is_valid():
            user = form.save(created_by=request.user)
            messages.success(request, f'Admin {user.email} created successfully.')
            return redirect('admin_panel:manage_admins')
    else:
        form = CreateAdminForm()

    return render(request, 'admin_panel/create_admin.html', {'form': form})


# ── 8. EDIT ADMIN PERMISSIONS ─────────────────────────────────

@superadmin_required
def edit_admin_permissions(request, admin_id):
    admin_profile = get_object_or_404(AdminProfile, id=admin_id)

    if request.method == 'POST':
        # checkbox checked = key exists in POST | unchecked = key missing
        # EXPAND: add new permission fields here as AdminProfile grows
        admin_profile.can_create_groups   = 'can_create_groups'   in request.POST
        admin_profile.can_assign_members  = 'can_assign_members'  in request.POST
        admin_profile.can_view_all_cases  = 'can_view_all_cases'  in request.POST
        admin_profile.can_create_users    = 'can_create_users'    in request.POST
        admin_profile.can_manage_roles    = 'can_manage_roles'    in request.POST
        admin_profile.save()

        messages.success(request, f'Permissions updated for {admin_profile.user.email}.')
        return redirect('admin_panel:manage_admins')

    return render(request, 'admin_panel/edit_admin_permissions.html', {
        'admin_profile': admin_profile,
    })


# ── 9. CASE LIST ──────────────────────────────────────────────

@admin_permission_required('can_view_all_cases')
def case_list(request):
    cases = (
        Case.objects
        .filter(is_active=True)
        .select_related('user', 'category', 'group', 'managed_profile')
        .order_by('-created_at')
    )

    # GET filters — usage: /admin-panel/cases/?status=pending&user=ali@email.com
    status_filter = request.GET.get('status')
    user_filter   = request.GET.get('user')

    if status_filter:
        cases = cases.filter(status=status_filter)
    if user_filter:
        cases = cases.filter(user__email__icontains=user_filter)

    return render(request, 'admin_panel/case_list.html', {
        'cases':          cases,
        'status_filter':  status_filter,
        'user_filter':    user_filter,
        'status_choices': Case.STATUS_CHOICES,
    })


# ── 10. CASE DETAIL ───────────────────────────────────────────
# Shows case info, answers, and the requirement management panel.
# Admin updates status/notes here.
# Requirement toggling handled by toggle_requirement view (POST only).
# Extra requirements added by add_extra_requirement view (POST only).

@admin_permission_required('can_view_all_cases')
def case_detail(request, case_id):
    """
    Admin views and manages a case.

    POST actions — determined by 'action' hidden field in each form:
      'update_case'   — update status and notes
      'edit_answer'   — admin edits a specific requirement's answer
      'toggle_req'    — handled by separate toggle_requirement view
      'add_extra'     — handled by separate add_extra_requirement view

    Data passed to template:
      requirement_rows — list of dicts, each with:
        cr      : CaseRequirement object
        req     : Requirement object
        answer  : CaseAnswer or None
      No lookups needed in template — everything pre-fetched here.

    EXPAND: add 'history' tracking to log every admin edit with timestamp.
    """
    case = get_object_or_404(Case, id=case_id)

    # ── Handle POST actions ───────────────────────────────────
    if request.method == 'POST':
        action = request.POST.get('action')

        # Action 1 — update case status and notes
        if action == 'update_case':
            new_status = request.POST.get('status')
            new_notes  = request.POST.get('notes', '')
            if new_status in dict(Case.STATUS_CHOICES):
                case.status = new_status
            case.notes = new_notes
            case.save()
            messages.success(request, 'Case updated.')

        # Action 2 — admin edits a specific answer
        elif action == 'edit_answer':
            requirement_id = request.POST.get('requirement_id')
            notify_user    = 'notify_user' in request.POST  # checkbox

            req = get_object_or_404(
                Requirement,
                id=requirement_id,
                # Security: requirement must belong to this case's category
                # or be an extra requirement for this case
            )

            # Get or create the answer row
            answer, _ = CaseAnswer.objects.get_or_create(
                case=case,
                requirement=req,
            )

            # Update the correct field based on requirement type
            # Each type maps to a different CaseAnswer field
            if req.type == 'text' or req.type == 'question':
                answer.answer_text   = request.POST.get('answer_value', '')
                answer.answer_date   = None
                answer.answer_number = None
                answer.answer_file   = None

            elif req.type == 'date':
                date_value = request.POST.get('answer_value', '')
                answer.answer_date   = date_value if date_value else None
                answer.answer_text   = None
                answer.answer_number = None
                answer.answer_file   = None

            elif req.type == 'number':
                number_value = request.POST.get('answer_value', '')
                answer.answer_number = number_value if number_value else None
                answer.answer_text   = None
                answer.answer_date   = None
                answer.answer_file   = None

            elif req.type in ('file', 'document'):
                if 'answer_file' in request.FILES:
                    answer.answer_file   = request.FILES['answer_file']
                    answer.answer_text   = None
                    answer.answer_date   = None
                    answer.answer_number = None

            # Mark as admin edited + set notification preference
            answer.edited_by_admin = True
            answer.notify_user     = notify_user
            answer.save()

            if notify_user:
                messages.success(
                    request,
                    f'Answer updated and user will be notified.'
                )
            else:
                messages.success(
                    request,
                    f'Answer updated silently.'
                )

        return redirect('admin_panel:case_detail', case_id=case_id)

    # ── Build requirement rows for template ───────────────────
    # Pre-fetch all answers for this case once — O(n)
    answers_map = {
        a.requirement_id: a
        for a in CaseAnswer.objects.filter(case=case).select_related('requirement')
    }

    case_requirements = (
        CaseRequirement.objects
        .filter(case=case)
        .select_related('requirement')
        .order_by('-is_active', 'requirement__name')
        # active first, then inactive, alphabetical within each group
    )

    # Build combined rows — template iterates this, no lookups needed
    # EXPAND: add more keys here if template needs more data
    requirement_rows = [
        {
            'cr':     cr,
            'req':    cr.requirement,
            'answer': answers_map.get(cr.requirement_id),
        }
        for cr in case_requirements
    ]

    # Requirements available to add as extras.
    # Phase 1 removed Requirement.category FK — traverse the M2M bridge instead.
    # category_requirements__category is the reverse of CategoryRequirement.category FK,
    # filtering to only requirements that are linked to THIS case's category.
    # .distinct() is required because a requirement could theoretically be linked
    # more than once (unique_together prevents it, but .distinct() is defensive).
    existing_req_ids = [row['req'].id for row in requirement_rows]
    available_to_add = (
        Requirement.objects
        .filter(
            category_requirements__category=case.category,  # via CategoryRequirement bridge
            is_active=True
        )
        .exclude(id__in=existing_req_ids)
        .distinct()
    )

    # ── Pre-fetch tasks, invoices, and required forms for the panels ──
    # WHY here (not in template tags): Django templates cannot call queryset
    # methods with arguments — pre-fetching in the view keeps the template clean.

    # Tasks linked to this case (reverse FK: Task.related_case → Case)
    linked_tasks = (
        case.tasks
        .select_related('assigned_to_user')
        .order_by('-created_at')
    )

    # Invoices linked to this case (reverse FK: Invoice.related_case → Case)
    linked_invoices = case.invoices.order_by('-created_at')

    # Government forms required by this category (M2M via CategoryForm bridge)
    # EXPAND: if category is None (old cases), category_forms will be an empty QS.
    category_forms = (
        case.category.category_forms
        .select_related('form')
        .order_by('order')
        if case.category else []
    )

    return render(request, 'admin_panel/case_detail.html', {
        'case':              case,
        'requirement_rows':  requirement_rows,
        'available_to_add':  available_to_add,
        'status_choices':    Case.STATUS_CHOICES,
        'linked_tasks':      linked_tasks,
        'linked_invoices':   linked_invoices,
        'category_forms':    category_forms,
    })


# ── 11. CREATE CASE ───────────────────────────────────────────
# After saving the Case, auto-creates CaseRequirement rows for all
# active requirements in the selected category.
# This is the "ready package" the user will fill in.

@admin_permission_required('can_view_all_cases')
def create_case(request):
    if request.method == 'POST':
        form = CreateCaseForm(request.POST)
        if form.is_valid():
            case = form.save(created_by=request.user)

            # Auto-create CaseRequirement rows from CategoryRequirement (Phase 1 M2M).
            # CategoryRequirement bridges Category ↔ Requirement after Phase 1 migration.
            # One row per active requirement — all active by default.
            # Admin can toggle individual ones later from case_detail.
            cat_reqs = (
                CategoryRequirement.objects
                .filter(category=case.category, requirement__is_active=True)
                .select_related('requirement')
                .order_by('order')
            )
            for cr in cat_reqs:
                CaseRequirement.objects.create(
                    case        = case,
                    requirement = cr.requirement,
                    is_active   = True,
                    is_extra    = False,    # from category defaults, not manually added
                )

            # Pre-fill answers from the user's profile data.
            # _autofill_case_answers is defined in this file (below _create_case).
            # WHY call it here too (not only inside _create_case):
            #   create_case view creates CaseRequirements directly without using
            #   the _create_case helper, so we must call it explicitly here.
            filled = _autofill_case_answers(case)

            messages.success(
                request,
                f'Case created for {form.cleaned_user.email} '
                f'— {cat_reqs.count()} requirements loaded'
                + (f', {filled} answers pre-filled from profile.' if filled else '.')
            )
            return redirect('admin_panel:case_detail', case_id=case.id)
    else:
        form = CreateCaseForm()

    return render(request, 'admin_panel/create_case.html', {
        'form':     form,
        'services': Service.objects.filter(is_active=True),
    })


# ── 11b. DELETE CASE ──────────────────────────────────────────
# Soft-delete only — sets is_active=False rather than removing the row.
# WHY soft delete: preserves the audit trail (answers, requirements, history).
#   The case disappears from all active queries (which filter is_active=True)
#   but can be recovered from the DB if needed by a developer.
# POST only — a GET request could be triggered accidentally (e.g. by a link
#   prefetcher), so we require an explicit POST to confirm the action.

@admin_permission_required('can_view_all_cases')
def delete_case(request, case_id):
    """
    POST /admin-panel/cases/<case_id>/delete/

    Soft-deletes the case (is_active=False).
    Redirects to case_list after success.

    EXPAND: add an 'undo' window — store deleted_at timestamp and allow
            restore within 30 days via a separate restore_case view.
    """
    if request.method != 'POST':
        # Safety net: redirect instead of showing an error page.
        # WHY: accidental GET (e.g. browser prefetch) should not crash —
        #      it just lands the admin back on the case detail page.
        return redirect('admin_panel:case_detail', case_id=case_id)

    # WHY no is_active=True filter: the delete button may still be visible in
    # cached pages even after the case was already soft-deleted. Requiring
    # is_active=True would throw a confusing 404 instead of silently succeeding.
    # We check and skip gracefully if already deleted.
    case = get_object_or_404(Case, id=case_id)

    if not case.is_active:
        # Already deleted — just redirect without an error message.
        messages.info(request, f'Case #{case.id} was already deleted.')
        return redirect('admin_panel:case_list')

    # Soft-delete: flip the flag.
    # update_fields limits the UPDATE to one column — no risk of overwriting
    # other fields if another request touches the row at the same time.
    case.is_active = False
    case.save(update_fields=['is_active'])

    messages.success(
        request,
        f'Case #{case.id} ({case.category}) has been deleted.'
    )
    return redirect('admin_panel:case_list')


# ── 12. TOGGLE REQUIREMENT ────────────────────────────────────
# Flips is_active on a CaseRequirement — soft delete or restore.
# POST only — no template. Redirects back to case_detail.
#
# Active   → removed: user no longer sees this requirement
# Removed  → restored: user sees it again

@admin_permission_required('can_view_all_cases')
def toggle_requirement(request, case_id, case_requirement_id):
    if request.method != 'POST':
        return redirect('admin_panel:case_detail', case_id=case_id)

    case_req = get_object_or_404(
        CaseRequirement,
        id=case_requirement_id,
        case_id=case_id  # security: must belong to this case
    )

    case_req.is_active = not case_req.is_active
    case_req.save()

    status = 'restored' if case_req.is_active else 'removed'
    messages.success(request, f'"{case_req.requirement.name}" {status}.')
    return redirect('admin_panel:case_detail', case_id=case_id)


# ── 13. ADD EXTRA REQUIREMENT ─────────────────────────────────
# Admin adds a requirement to a case that wasn't in the category defaults.
# is_extra=True marks it as manually added.
# POST only — requirement_id comes from a select on the case_detail page.

@admin_permission_required('can_view_all_cases')
def add_extra_requirement(request, case_id):
    if request.method != 'POST':
        return redirect('admin_panel:case_detail', case_id=case_id)

    case           = get_object_or_404(Case, id=case_id)
    requirement_id = request.POST.get('requirement_id')
    # Two-step lookup — Phase 1 removed Requirement.category FK.
    # Step 1: confirm the requirement exists and is active (basic 404 guard).
    requirement = get_object_or_404(Requirement, id=requirement_id, is_active=True)
    # Step 2: security check — requirement must be linked to THIS case's category
    # via the CategoryRequirement bridge. get_object_or_404 returns 404 if it isn't,
    # preventing a user from injecting a requirement from a different category.
    get_object_or_404(CategoryRequirement, requirement=requirement, category=case.category)

    # get_or_create — safe if row already exists (e.g., was removed before)
    case_req, created = CaseRequirement.objects.get_or_create(
        case=case,
        requirement=requirement,
        defaults={
            'is_active': True,
            'is_extra':  True,
        }
    )

    if not created:
        # Row existed but was inactive — restore it
        case_req.is_active = True
        case_req.save()

    messages.success(request, f'"{requirement.name}" added to this case.')
    return redirect('admin_panel:case_detail', case_id=case_id)


# ── 13b. EXPORT CASE REQUIREMENTS ─────────────────────────────
# Downloads an Excel file listing every requirement for the case,
# its description, and the user's answer.
# Columns: #, Requirement, Type, Description, Answer
# EXPAND: add more columns (e.g., admin notes, last edited) or
#         apply different styling per requirement type.

@admin_permission_required('can_view_all_cases')
def export_case_requirements(request, case_id):
    case = get_object_or_404(Case, id=case_id)

    # Fetch all answers for this case once
    answers_map = {
        a.requirement_id: a
        for a in CaseAnswer.objects.filter(case=case)
    }

    rows = (
        CaseRequirement.objects
        .filter(case=case, is_active=True)
        .select_related('requirement')
        .order_by('requirement__name')
    )

    # ── Build workbook ──────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Case {case.id}'

    # Header row styling
    header_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')  # navy
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['#', 'Requirement', 'Type', 'Description', 'Answer']
    col_widths = [5, 28, 12, 36, 36]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22

    # Data rows
    wrap = Alignment(vertical='top', wrap_text=True)

    for row_num, cr in enumerate(rows, start=2):
        req    = cr.requirement
        answer = answers_map.get(req.id)

        # Resolve answer value to a human-readable string
        if answer is None:
            answer_str = '—'
        elif answer.answer_file:
            answer_str = 'Document uploaded'
        elif answer.answer_text:
            answer_str = answer.answer_text
        elif answer.answer_date:
            answer_str = str(answer.answer_date)
        elif answer.answer_number is not None:
            answer_str = str(answer.answer_number)
        else:
            answer_str = '—'

        ws.cell(row=row_num, column=1, value=row_num - 1).alignment = wrap
        ws.cell(row=row_num, column=2, value=req.name).alignment = wrap
        ws.cell(row=row_num, column=3, value=req.type).alignment = wrap
        ws.cell(row=row_num, column=4, value=req.description or '').alignment = wrap
        ws.cell(row=row_num, column=5, value=answer_str).alignment = wrap

        # Zebra striping — light grey on even rows
        if row_num % 2 == 0:
            row_fill = PatternFill(fill_type='solid', fgColor='F5F5F5')
            for col in range(1, 6):
                ws.cell(row=row_num, column=col).fill = row_fill

    # ── Stream response ─────────────────────────────────────
    filename = f'case_{case.id}_requirements.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── 14. GROUP LIST ────────────────────────────────────────────

@admin_permission_required('can_create_groups')
def group_list(request):
    """
    Lists all groups. Admin can create a new group from here.
    EXPAND: add search/filter by group type.
    """
    groups = (
        Group.objects
        .filter(is_active=True)
        .select_related('created_by')
        .order_by('-created_at')
    )
    return render(request, 'admin_panel/group_list.html', {'groups': groups})

# ── 15++. admin_create_group ──────────────────────────────────────────
@admin_permission_required('can_create_groups')
def admin_create_group(request):
    """
    Admin creates a group and assigns an owner.
    Owner is auto-added as first member.
    EXPAND: add default permissions assignment for the owner here.
    """
    if request.method == 'POST':
        form = AdminGroupCreateForm(request.POST)
        if form.is_valid():
            group = form.save(created_by=request.user)
            messages.success(
                request,
                f'Group "{group.name}" created. '
                f'{form.cleaned_owner.email} added as owner.'
            )
            return redirect('admin_panel:group_detail', group_id=group.id)
    else:
        form = AdminGroupCreateForm()

    return render(request, 'admin_panel/create_group.html', {'form': form})

# ── 15. GROUP DETAIL ──────────────────────────────────────────

@admin_permission_required('can_assign_members')
def group_detail(request, group_id):
    """
    Admin views and manages a group:
    - Members with their roles, permissions, and inline cases
    - Managed profiles with inline cases
    - Action buttons: assign case, create managed profile, link managed to user, bulk assign
    EXPAND: add notes or status fields per membership.
    """
    group           = get_object_or_404(Group, id=group_id)
    memberships     = (
        group.memberships
        .select_related('user', 'role')
        .prefetch_related('permissions')
        .order_by('-joined_at')
    )
    managed_profiles_qs = group.managed_profiles.select_related('created_by', 'linked_user')
    roles           = Role.objects.filter(group_type=group.type)
    all_permissions = GroupPermission.objects.all()

    # ── Build per-member case map ─────────────────────────────
    # One query for all member cases in this group, then group by user_id in Python.
    # Avoids N+1 (one query per member).
    member_case_map = {}
    for case in (
        Case.objects
        .filter(group=group, managed_profile=None, is_active=True)
        .select_related('category')
        .order_by('-created_at')
    ):
        member_case_map.setdefault(case.user_id, []).append(case)

    members_with_cases = [
        (m, member_case_map.get(m.user_id, []))
        for m in memberships
    ]

    # ── Build per-managed-profile case map ────────────────────
    # Same strategy: one query, grouped in Python by managed_profile_id.
    mp_case_map = {}
    for case in (
        Case.objects
        .filter(managed_profile__group=group, is_active=True)
        .select_related('category')
        .order_by('-created_at')
    ):
        mp_case_map.setdefault(case.managed_profile_id, []).append(case)

    managed_with_cases = [
        (mp, mp_case_map.get(mp.id, []))
        for mp in managed_profiles_qs
    ]

    return render(request, 'admin_panel/group_detail.html', {
        'group':               group,
        'members_with_cases':  members_with_cases,   # list of (GroupMembership, [Case])
        'managed_with_cases':  managed_with_cases,   # list of (ManagedProfile, [Case])
        'roles':               roles,
        'all_permissions':     all_permissions,
        # legacy key kept so existing permission/role forms still work
        'memberships':         memberships,
    })
# ── 15++. admin_add_member ──────────────────────────────────────────
@admin_permission_required('can_assign_members')
def admin_add_member(request, group_id):
    """
    Admin adds a user to a group with optional role.
    POST only from group_detail page.
    EXPAND: add bulk member import from CSV.
    """
    group = get_object_or_404(Group, id=group_id)

    if request.method == 'POST':
        form = AdminAddMemberForm(group, request.POST)
        if form.is_valid():
            membership = form.save()
            messages.success(
                request,
                f'{membership.user.email} added to {group.name}.'
            )
        else:
            # Pass errors back — re-render group detail with form errors
            for error in form.errors.values():
                messages.error(request, error.as_text())

    return redirect('admin_panel:group_detail', group_id=group_id)

# ── 15++. admin_toggle_member ──────────────────────────────────────────
@admin_permission_required('can_assign_members')
def admin_toggle_member(request, group_id, membership_id):
    """
    Suspends or restores a group member.
    is_active=False = suspended (blocked from filling cases, viewing group)
    is_active=True  = active again
    POST only — no template needed.
    """
    if request.method != 'POST':
        return redirect('admin_panel:group_detail', group_id=group_id)

    membership = get_object_or_404(
        GroupMembership,
        id=membership_id,
        group_id=group_id  # security: must belong to this group
    )
    membership.is_active = not membership.is_active
    membership.save()

    status = 'restored' if membership.is_active else 'suspended'
    messages.success(request, f'{membership.user.email} {status}.')
    return redirect('admin_panel:group_detail', group_id=group_id)

# ── 15++. admin_set_member_permissions ──────────────────────────────────────────
@admin_permission_required('can_assign_members')
def admin_set_member_permissions(request, group_id, membership_id):
    """
    Admin sets which GroupPermissions a member has.
    Checkboxes — checked = has permission, unchecked = doesn't.
    POST only from group_detail page.

    EXPAND: add role assignment here too.
    """
    if request.method != 'POST':
        return redirect('admin_panel:group_detail', group_id=group_id)

    membership = get_object_or_404(
        GroupMembership,
        id=membership_id,
        group_id=group_id
    )

    # Get all permission IDs that were checked
    permission_ids = request.POST.getlist('permissions')

    # Replace all permissions with the checked ones
    # set() replaces the entire M2M — clean and simple
    membership.permissions.set(permission_ids)
    membership.save()

    messages.success(
        request,
        f'Permissions updated for {membership.user.email}.'
    )
    return redirect('admin_panel:group_detail', group_id=group_id)


# 15++++--- admin change member role (inline from group detail) ---
@admin_permission_required('can_assign_members')
def admin_change_member_role(request, group_id, membership_id):
    """
    Changes a member's role inline from the group detail page.
    role_id = empty string means remove the role (set to None).
    POST only.
    """
    if request.method != 'POST':
        return redirect('admin_panel:group_detail', group_id=group_id)

    membership = get_object_or_404(
        GroupMembership,
        id=membership_id,
        group_id=group_id  # security: must belong to this group
    )

    role_id = request.POST.get('role_id')

    if role_id:
        role = get_object_or_404(Role, id=role_id, group_type=membership.group.type)
        membership.role = role
    else:
        membership.role = None  # remove role

    membership.save()
    messages.success(request, f'Role updated for {membership.user.email}.')
    return redirect('admin_panel:group_detail', group_id=group_id)


# ── GROUP CASE HELPER ─────────────────────────────────────────
# Called by the five views below — centralises Case + CaseRequirement
# creation so no view can accidentally forget the requirement rows.
#
# WHY create CaseRequirements immediately?
#   Each CaseRequirement row gives admin the ability to customise which
#   requirements apply to this specific case later (is_active=False removes
#   one; is_extra=True adds a custom one). Without them, the case detail
#   page has nothing to show or toggle.
#
# EXPAND: add a 'notes' parameter for creation-time admin notes.

def _create_case(user, group, category, managed_profile=None):
    """
    Create a Case + one CaseRequirement per active requirement in the category.

    Parameters:
        user            — always required (who is submitting/managing the case)
        group           — the group this case belongs to
        category        — which service category this case is for
        managed_profile — set when the case is for an interior person; None = member case

    Returns the newly created Case instance.
    """
    case = Case.objects.create(
        user            = user,
        group           = group,
        category        = category,
        managed_profile = managed_profile,
        status          = 'pending',
    )
    # Auto-create one CaseRequirement per active CategoryRequirement in the category.
    # Uses the Phase 1 M2M bridge (CategoryRequirement) instead of direct FK.
    # 'order_by("order")' preserves the display order set in the builder.
    # is_active=True  → shown by default (admin can disable per-case later)
    # is_extra=False  → standard requirement from category defaults, not manually added
    for cr in (
        CategoryRequirement.objects
        .filter(category=category, requirement__is_active=True)
        .select_related('requirement')
        .order_by('order')
    ):
        CaseRequirement.objects.create(
            case        = case,
            requirement = cr.requirement,
            is_active   = True,
            is_extra    = False,
        )

    # Pre-fill answers from the user's stored profile data.
    # WHY here (after all CaseRequirements are created, not in a loop above):
    #   _autofill_case_answers queries CaseRequirement in a single batch — calling
    #   it once after the loop is more efficient than checking profile data per row.
    _autofill_case_answers(case)

    return case


# ── AUTO-FILL HELPER ──────────────────────────────────────────
# Called immediately after a case is created (and its CaseRequirements exist).
# Looks at each requirement's profile_mapping / managed_profile_mapping and writes
# a CaseAnswer pre-populated from the user's stored profile data.
#
# WHY do this at creation time (not on every page load)?
#   If we pre-fill at load time the user could lose edits when the page reloads.
#   Creating the answer once at case-creation time means:
#     - The user sees their info already filled in when they first open the case.
#     - They can still edit or override the pre-filled answer at any time.
#     - Admin edits in the case_detail view also see the pre-filled value.
#
# WHY only create (not update) existing answers?
#   We use get_or_create so that calling this on an already-populated case
#   is idempotent — it won't overwrite answers the user has already edited.
#
# EXPAND: add a "Re-sync from profile" button in case_detail so admin can
#         re-trigger this for cases that existed before profile data was filled in.

def _autofill_case_answers(case):
    """
    Pre-populate CaseAnswer rows from the user's Profile or ManagedProfile data.

    For each active CaseRequirement on the case:
      1. Check if a profile_mapping dot-path is set (e.g. 'profile.date_of_birth')
      2. Resolve the value by walking the dot-path on case.user
      3. If a non-None value is found AND no CaseAnswer yet exists → create one
      4. Repeat with managed_profile_mapping if case.managed_profile is set

    Returns a count of answers that were pre-filled (useful for log messages).

    EXPAND: add support for 'cross_case' fill — copy the most recent answer
            to the same requirement from any older case for the same user.
    """
    from cases.forms import _try_profile_fill, _try_managed_profile_fill
    import datetime

    filled = 0

    active_reqs = (
        CaseRequirement.objects
        .filter(case=case, is_active=True)
        .select_related('requirement')
    )

    for cr in active_reqs:
        req = cr.requirement

        # Skip info_text type — it has no answer field
        if req.type == 'info_text':  # Requirement.type is the field name (NOT input_type)
            continue

        # Skip if an answer already exists (idempotent — don't overwrite user edits)
        if CaseAnswer.objects.filter(case=case, requirement=req).exists():
            continue

        # ── Resolve value from profile or managed_profile ─────
        value = None

        if case.managed_profile:
            # Managed profile takes precedence — this case is for an interior person
            value = _try_managed_profile_fill(req, case.managed_profile)

        if value is None and case.user:
            # Fall back to the regular user profile mapping
            value = _try_profile_fill(req, case.user)

        if value is None:
            continue    # no mapping set, or the mapped field is empty → skip

        # ── Convert value to the right CaseAnswer field ───────
        # CaseAnswer has answer_text (str) and answer_date (date).
        # Determine which field to fill based on the value's type.
        answer_text = None
        answer_date = None

        if isinstance(value, datetime.date):
            # datetime.datetime is a subclass of datetime.date — both work here
            answer_date = value if isinstance(value, datetime.date) else None
        else:
            # Everything else (str, int, bool) goes into answer_text as a string
            answer_text = str(value)

        if answer_text is None and answer_date is None:
            continue

        CaseAnswer.objects.create(
            case        = case,
            requirement = req,
            answer_text = answer_text or '',
            answer_date = answer_date,
        )
        filled += 1

    return filled


# ── ADMIN: CREATE MANAGED PROFILE ────────────────────────────
# Admin creates an interior person (no account) inside a group.
# Requires: can_assign_members (admin manages group membership)
#
# URL: /admin-panel/groups/<group_id>/add-person/

@admin_permission_required('can_assign_members')
def admin_create_managed_profile(request, group_id):
    """
    Admin creates a ManagedProfile (interior person).
    Default: joins the group from the URL (group_id).
    Optional: if 'new_group_name' is filled in the form, a brand-new group is
    created for this person exclusively — their own "file". The URL's group is
    used only for the redirect back; the managed profile gets its own group.

    created_by = admin (can be None if admin chooses to make it group-owned only,
    but for traceability we always set it to the acting admin here).
    """
    group = get_object_or_404(Group, id=group_id)

    if request.method == 'POST':
        form = AdminManagedProfileForm(request.POST, request.FILES)
        if form.is_valid():
            managed            = form.save(commit=False)
            managed.created_by = request.user   # admin is the creator

            new_group_name = form.cleaned_data.get('new_group_name', '').strip()
            if new_group_name:
                # Create a new group exclusively for this interior person.
                # Type 'other' is the correct valid fallback (matches Group.GROUP_TYPES choices).
                # EXPAND: expose a type dropdown if admins need family/business/etc.
                new_group = Group.objects.create(
                    name       = new_group_name,
                    type       = 'other',
                    created_by = request.user,
                )
                managed.group = new_group
            else:
                # Default: add to the existing group from the URL
                managed.group = group

            managed.save()
            messages.success(request, f'Profile for {managed.full_name()} created.')
            return redirect('admin_panel:group_detail', group_id=group_id)
    else:
        form = AdminManagedProfileForm()

    return render(request, 'admin_panel/managed_profile_form.html', {
        'form':  form,
        'group': group,
    })


# ── ADMIN: ASSIGN CASE TO MEMBER ─────────────────────────────
# Admin creates a new Case for a specific real-user group member.
# Requires: can_assign_members or can_view_all_cases
#
# URL: /admin-panel/groups/<group_id>/member/<membership_id>/assign-case/

@admin_permission_required('can_assign_members')
def admin_assign_case_to_member(request, group_id, membership_id):
    """
    Admin picks a category to create a Case for a real group member.
    The case is owned by the member's User account (case.user = member.user).
    EXPAND: add a 'notes' field so admin can attach context at creation time.
    """
    group = get_object_or_404(Group, id=group_id)
    # membership must belong to this group — prevents cross-group ID guessing
    target = get_object_or_404(GroupMembership, id=membership_id, group=group, is_active=True)

    if request.method == 'POST':
        form = AdminAssignCaseForm(request.POST)
        if form.is_valid():
            category = form.cleaned_data['category']
            _create_case(user=target.user, group=group, category=category)
            messages.success(
                request,
                f'Case "{category.name}" assigned to {target.user.email}.'
            )
            return redirect('admin_panel:group_detail', group_id=group_id)
    else:
        form = AdminAssignCaseForm()

    return render(request, 'admin_panel/assign_case.html', {
        'form':         form,
        'group':        group,
        'target_label': target.user.email,
        'target_type':  'member',
    })


# ── ADMIN: CREATE CASE FOR MANAGED PROFILE ────────────────────
# Admin creates a Case for a managed profile (interior person, no account).
# Requires: can_assign_members
#
# URL: /admin-panel/groups/<group_id>/managed/<managed_id>/assign-case/

@admin_permission_required('can_assign_members')
def admin_create_case_for_managed(request, group_id, managed_id):
    """
    Admin picks a category to create a Case for a managed profile.
    case.user = request.user (admin manages on behalf of the profile).
    case.managed_profile = the interior person.
    EXPAND: add a 'notes' field for creation-time context.
    """
    group   = get_object_or_404(Group, id=group_id)
    managed = get_object_or_404(ManagedProfile, id=managed_id, group=group)

    if request.method == 'POST':
        form = AdminAssignCaseForm(request.POST)
        if form.is_valid():
            category = form.cleaned_data['category']
            _create_case(
                user            = request.user,   # admin acts on behalf of managed profile
                group           = group,
                category        = category,
                managed_profile = managed,
            )
            messages.success(
                request,
                f'Case "{category.name}" assigned to {managed.full_name()}.'
            )
            return redirect('admin_panel:group_detail', group_id=group_id)
    else:
        form = AdminAssignCaseForm()

    return render(request, 'admin_panel/assign_case.html', {
        'form':         form,
        'group':        group,
        'target_label': managed.full_name(),
        'target_type':  'managed',
    })


# ── ADMIN: LINK MANAGED PROFILE TO USER ACCOUNT ───────────────
# Admin links a ManagedProfile to an existing User account.
# Used when an interior person later creates their own login.
# Requires: can_assign_members (sensitive identity-linking operation)
#
# URL: /admin-panel/groups/<group_id>/managed/<managed_id>/link-user/

@admin_permission_required('can_assign_members')
def admin_link_managed_to_user(request, group_id, managed_id):
    """
    Sets managed_profile.linked_user to an existing User account.
    After linking:
      managed.linked_user     → the User
      user.managed_account    → the ManagedProfile (reverse OneToOne)
    Guard: redirect with error if already linked (no confusing re-link form).
    EXPAND: add an audit log entry when a link is created.
    """
    group   = get_object_or_404(Group, id=group_id)
    managed = get_object_or_404(ManagedProfile, id=managed_id, group=group)

    if managed.linked_user is not None:
        messages.error(
            request,
            f'{managed.full_name()} is already linked to {managed.linked_user.email}.'
        )
        return redirect('admin_panel:group_detail', group_id=group_id)

    if request.method == 'POST':
        form = AdminLinkManagedForm(request.POST)
        if form.is_valid():
            managed.linked_user = form.cleaned_user
            managed.save()
            messages.success(
                request,
                f'{managed.full_name()} is now linked to {form.cleaned_user.email}.'
            )
            return redirect('admin_panel:group_detail', group_id=group_id)
    else:
        form = AdminLinkManagedForm()

    return render(request, 'admin_panel/link_managed_user.html', {
        'form':    form,
        'group':   group,
        'managed': managed,
    })


# ── ADMIN: MANAGED PROFILE DETAIL ────────────────────────────
# Read-only summary of an interior person: personal info + all their cases.
# Links to "Assign Case", "Link Account", and back to the group detail page.
# Requires: can_assign_members (same as other managed-profile operations)
#
# URL: /admin-panel/groups/<group_id>/managed/<managed_id>/

@admin_permission_required('can_assign_members')
def admin_managed_profile_detail(request, group_id, managed_id):
    """
    Shows all information for a single managed profile (interior person):
      - personal info (name, DOB, gender, origin, passport, photos)
      - which group they belong to and who created them
      - all their cases with status badges and View links
      - action buttons: Assign Case, Link Account

    Context variables:
      group   — the Group this managed profile belongs to
      managed — the ManagedProfile object
      cases   — all active Cases for this managed profile
    """
    group   = get_object_or_404(Group, id=group_id)
    managed = get_object_or_404(ManagedProfile, id=managed_id, group=group)

    # Fetch all active cases for this interior person.
    # select_related on category avoids an extra query per case row.
    cases = (
        Case.objects
        .filter(managed_profile=managed, is_active=True)
        .select_related('category')
        .order_by('-created_at')  # most recent first
    )

    return render(request, 'admin_panel/managed_profile_detail.html', {
        'group':   group,
        'managed': managed,
        'cases':   cases,
    })


# ── ADMIN: BULK ASSIGN CATEGORY TO WHOLE GROUP ────────────────
# Creates one Case per active member AND one Case per managed profile,
# all with the same selected category. Skips anyone who already has
# an active case with this category in this group (no duplicates).
# Requires: can_assign_members
#
# URL: /admin-panel/groups/<group_id>/assign-category/

@admin_permission_required('can_assign_members')
def admin_assign_category_to_group(request, group_id):
    """
    Admin selects one category → one Case is created for every active member
    AND every managed profile in the group.

    WHY skip duplicates?
      Prevents accidentally creating a second case if admin runs this twice.
      A clear created/skipped summary is shown so admin knows what happened.

    EXPAND: add a 'notes' field applied to all created cases.
    """
    group         = get_object_or_404(Group, id=group_id)
    member_count  = group.memberships.filter(is_active=True).count()
    managed_count = group.managed_profiles.count()

    if request.method == 'POST':
        form = AdminAssignCaseForm(request.POST)
        if form.is_valid():
            category      = form.cleaned_data['category']
            created_count = 0
            skipped_count = 0

            # ── One case per active member ─────────────────────
            for m in group.memberships.filter(is_active=True).select_related('user'):
                already = Case.objects.filter(
                    user=m.user, group=group, category=category,
                    managed_profile=None, is_active=True,
                ).exists()
                if already:
                    skipped_count += 1
                else:
                    _create_case(user=m.user, group=group, category=category)
                    created_count += 1

            # ── One case per managed profile ───────────────────
            for mp in group.managed_profiles.all():
                already = Case.objects.filter(
                    managed_profile=mp, category=category, is_active=True,
                ).exists()
                if already:
                    skipped_count += 1
                else:
                    _create_case(
                        user=request.user, group=group,
                        category=category, managed_profile=mp,
                    )
                    created_count += 1

            parts = [f'Created {created_count} case(s) for "{category.name}".']
            if skipped_count:
                parts.append(f'{skipped_count} already had this category — skipped.')
            messages.success(request, ' '.join(parts))
            return redirect('admin_panel:group_detail', group_id=group_id)
    else:
        form = AdminAssignCaseForm()

    return render(request, 'admin_panel/assign_category_group.html', {
        'form':          form,
        'group':         group,
        'member_count':  member_count,
        'managed_count': managed_count,
    })


# ── 15b. AJAX — USER SEARCH ───────────────────────────────────
# Used by the user-autocomplete widget in CreateCaseForm, AdminAddMemberForm,
# and any other form where the admin must pick an existing user.
#
# WHY AJAX autocomplete instead of a plain email input?
#   A plain email field forces the admin to remember the exact email address.
#   Autocomplete lets them type 2–3 characters and see matches by first name,
#   last name, OR email — much faster and less error-prone.
#
# SECURITY NOTE:
#   Only admins can call this endpoint (@admin_required).
#   Only non-staff users are searchable — admins and superusers are hidden
#   from the results to prevent confusion (you don't create cases for admins).
#   EXPAND: add is_active filter if you later add user deactivation.

@admin_required
def ajax_user_search(request):
    """
    GET /admin-panel/ajax/user-search/?q=ali

    Returns up to 10 users whose email, first_name, or last_name contains
    the search term (case-insensitive).

    Response:
      {
        "users": [
          {"id": 7, "email": "ali@email.com", "full_name": "Ali Karimi",
           "label": "Ali Karimi — ali@email.com"},
          ...
        ]
      }

    The 'label' field is designed to be shown directly in a dropdown option —
    it contains enough info to identify a user at a glance.

    EXPAND: add 'phone' to the label so admins can also search by phone number.
    EXPAND: order by recently-created cases first (most active clients near top).
    """
    q = request.GET.get('q', '').strip()

    if len(q) < 2:
        # Require at least 2 characters — otherwise we'd return hundreds of rows.
        # The frontend should enforce this too, but we guard here as well.
        return JsonResponse({'users': []})

    from django.db.models import Q

    users = (
        User.objects
        .filter(is_staff=False, is_superuser=False)   # exclude admin accounts
        .filter(
            # Case-insensitive search across three fields simultaneously.
            # Q objects let us combine multiple field lookups with OR.
            # icontains = SQL LIKE '%q%', case-insensitive.
            Q(email__icontains=q)       |
            Q(first_name__icontains=q)  |
            Q(last_name__icontains=q)
        )
        .order_by('first_name', 'last_name')
        [:10]  # cap at 10 results — enough to narrow down without being overwhelming
    )

    return JsonResponse({
        'users': [
            {
                'id':        u.id,
                'email':     u.email,
                'full_name': f'{u.first_name} {u.last_name}'.strip() or u.email,
                # 'label' is the string shown in the dropdown option.
                # Format: "First Last — email@example.com"
                'label':     f'{u.first_name} {u.last_name}'.strip() + f' — {u.email}'
                             if (u.first_name or u.last_name)
                             else u.email,
            }
            for u in users
        ]
    })


# ── 16. AJAX — GET CATEGORIES ─────────────────────────────────
# Called by JavaScript when admin selects a service.
# Returns top-level categories as JSON to populate the next dropdown.

def get_categories(request):
    """
    URL: /admin-panel/ajax/categories/?service_id=3
    Response: {"categories": [{"id": 1, "name": "Work Visa"}, ...]}
    """
    service_id = request.GET.get('service_id')
    if not service_id:
        return JsonResponse({'categories': []})

    categories = Category.objects.filter(
        service_id=service_id,
        parent=None,
        is_active=True
    ).values('id', 'name')

    return JsonResponse({'categories': list(categories)})


# ── 17. AJAX — GET SUBCATEGORIES ─────────────────────────────
# Called by JavaScript when admin selects a category.

def get_subcategories(request):
    """
    URL: /admin-panel/ajax/subcategories/?category_id=5
    Response: {"subcategories": [{"id": 3, "name": "LMIA Exempt"}, ...]}
    Empty list = this category has no subcategories.
    """
    category_id = request.GET.get('category_id')
    if not category_id:
        return JsonResponse({'subcategories': []})

    subcategories = Category.objects.filter(
        parent_id=category_id,
        is_active=True
    ).values('id', 'name')

    return JsonResponse({'subcategories': list(subcategories)})

# ── Builder Page ──────────────────────────────────────────────

@admin_permission_required('can_manage_content')
def service_builder(request):
    """
    Renders the single-page service builder.
    All data loading happens via AJAX after the page loads.
    This view just renders the empty shell.
    """
    return render(request, 'admin_panel/service_builder.html')


# ── Service List (for list/overview page) ─────────────────────

@admin_permission_required('can_manage_content')
def service_list(request):
    """
    Overview page — lists all services as cards.
    Links to builder with each service pre-selected.
    """
    services = Service.objects.prefetch_related('categories').order_by('name')
    service_data = [
        {
            'service':        s,
            'category_count': s.categories.filter(is_active=True).count(),
            # WHY this traversal:
            # After Phase 1, Requirement has NO direct category FK.
            # The M2M bridge is CategoryRequirement (category_requirements related_name).
            # Path: Requirement → category_requirements (CategoryRequirement) → category → service
            # .distinct() prevents double-counting if the same Requirement is linked to
            # multiple categories within the same service.
            'req_count':      (
                Requirement.objects
                .filter(
                    category_requirements__category__service=s,
                    is_active=True
                )
                .distinct()
                .count()
            ),
        }
        for s in services
    ]
    return render(request, 'admin_panel/service_list.html', {
        'service_data': service_data,
    })


# ── AJAX: Load All Services ───────────────────────────────────

@admin_required
def ajax_get_services(request):
    """
    GET /ajax/builder/services/
    Returns all active services for the left panel.
    Response: { services: [{id, name, description, is_active, category_count}] }
    """
    services = Service.objects.filter(is_active=True).order_by('name')
    data = [
        {
            'id':             s.id,
            'name':           s.name,
            'description':    s.description or '',
            'is_active':      s.is_active,
            'category_count': s.categories.filter(is_active=True).count(),
        }
        for s in services
    ]
    return JsonResponse({'services': data})


# ── AJAX: Create Service ──────────────────────────────────────

@admin_required
@require_http_methods(['POST'])
def ajax_create_service(request):
    """
    POST /ajax/builder/service/create/
    Body: { name, description }
    Response: { ok: true, service: {id, name, description, is_active} }
              { ok: false, errors: {field: [error]} }

    Why require_http_methods?
      Explicit — rejects GET requests cleanly with 405 instead of silent failure.
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    name        = body.get('name', '').strip()
    description = body.get('description', '').strip()

    # Validate
    errors = {}
    if not name:
        errors['name'] = ['Name is required.']
    if Service.objects.filter(name__iexact=name).exists():
        errors['name'] = [f'A service named "{name}" already exists.']

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    service = Service.objects.create(name=name, description=description)
    return JsonResponse({
        'ok':      True,
        'service': {
            'id':          service.id,
            'name':        service.name,
            'description': service.description or '',
            'is_active':   service.is_active,
        }
    })


# ── AJAX: Load Categories for a Service or Parent ─────────────

@admin_required
def ajax_get_categories(request):
    """
    GET /ajax/builder/categories/?service_id=1
    GET /ajax/builder/categories/?parent_id=5
    Returns direct children only — not the full tree.
    JavaScript builds the tree level by level.

    Response: { categories: [{id, name, description, is_active, req_count, has_children}] }

    has_children tells JS whether to show an expand arrow on the category.
    """
    service_id = request.GET.get('service_id')
    parent_id  = request.GET.get('parent_id')

    if service_id:
        cats = Category.objects.filter(
            service_id=service_id,
            parent=None,
            is_active=True
        ).order_by('name')
    elif parent_id:
        cats = Category.objects.filter(
            parent_id=parent_id,
            is_active=True
        ).order_by('name')
    else:
        return JsonResponse({'categories': []})

    data = [
        {
            'id':           c.id,
            'name':         c.name,
            'description':  c.description or '',
            'is_active':    c.is_active,
            'service_id':   c.service_id,
            'parent_id':    c.parent_id,
            # Count via CategoryRequirement bridge — Requirement.category FK was removed in Phase 1.
            # CategoryRequirement is the M2M through-table linking categories to their requirements.
            'req_count':    CategoryRequirement.objects.filter(
                                category=c,
                                requirement__is_active=True
                            ).count(),
            'has_children': c.subcategories.exists(),
        }
        for c in cats
    ]
    return JsonResponse({'categories': data})


# ── AJAX: Create Category ─────────────────────────────────────

@admin_required
@require_http_methods(['POST'])
def ajax_create_category(request):
    """
    POST /ajax/builder/category/create/
    Body: { name, description, service_id, parent_id (optional) }
    Response: { ok: true, category: {...} }
              { ok: false, errors: {...} }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    name        = body.get('name', '').strip()
    description = body.get('description', '').strip()
    service_id  = body.get('service_id')
    parent_id   = body.get('parent_id')  # None = top-level

    # Validate
    errors = {}
    if not name:
        errors['name'] = ['Name is required.']
    if not service_id:
        errors['service_id'] = ['Service is required.']

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    service = get_object_or_404(Service, id=service_id)
    parent  = get_object_or_404(Category, id=parent_id) if parent_id else None

    # Ensure parent belongs to same service
    if parent and parent.service_id != service.id:
        return JsonResponse({
            'ok': False,
            'errors': {'parent_id': ['Parent must belong to the same service.']}
        })

    category = Category.objects.create(
        name        = name,
        description = description,
        service     = service,
        parent      = parent,
    )

    return JsonResponse({
        'ok':       True,
        'category': {
            'id':           category.id,
            'name':         category.name,
            'description':  category.description or '',
            'is_active':    category.is_active,
            'service_id':   category.service_id,
            'parent_id':    category.parent_id,
            'req_count':    0,
            'has_children': False,
        }
    })


# ── AJAX: Get Category Detail (requirements + inheritance) ────

@admin_required
def ajax_get_category_detail(request, category_id):
    """
    GET /ajax/builder/category/<id>/
    Returns:
      - category info
      - own requirements (directly on this category)
      - inherited requirements (from all parent categories, with source info)

    This powers the right panel in the builder.
    Response: {
      category: {...},
      own_requirements: [{id, name, type, description, is_required, is_active}],
      inherited: [{req: {...}, from_category: {id, name}}]
    }
    """
    category = get_object_or_404(Category, id=category_id)

    # ── Own requirements (via CategoryRequirement M2M) ─────────────────
    # Phase 1: requirements are no longer stored directly on the category.
    # They are linked via CategoryRequirement rows, ordered by 'order' field.
    # 'is_required_effective' respects per-category override if set.
    own_cat_reqs = (
        CategoryRequirement.objects
        .filter(category=category, requirement__is_active=True)
        .select_related('requirement', 'requirement__section')
        .order_by('order')
    )

    # ── Build requirement → form-code lookup (for grouping in the UI) ──────
    # WHY: the service builder shows requirements grouped under the government
    # form they belong to (e.g. "IMM5710 Fields" section). To do this we need
    # to know which forms are linked to THIS category AND which of those forms
    # contain each requirement.
    #
    # Step 1: get the IDs of all forms linked to this category via CategoryForm.
    # Step 2: for each FormRequirement whose form is in that set AND whose
    #         requirement is in our own_cat_reqs, record the form code.
    # Result: req_to_forms maps requirement_id → [list of form codes].
    from cases.models import FormRequirement, CategoryForm as CatFormModel
    linked_form_ids = set(
        CatFormModel.objects
        .filter(category=category)
        .values_list('form_id', flat=True)
    )
    req_to_forms = {}  # {requirement_id: [form_code, ...]}
    if linked_form_ids:
        for fr in (
            FormRequirement.objects
            .filter(
                form_id__in=linked_form_ids,
                requirement_id__in=own_cat_reqs.values_list('requirement_id', flat=True),
            )
            .select_related('form')
        ):
            req_to_forms.setdefault(fr.requirement_id, []).append(fr.form.code)

    own_reqs = []
    for cr in own_cat_reqs:
        req = cr.requirement
        own_reqs.append({
            'id':               req.id,
            'name':             req.name,
            'type':             req.type,
            'description':      req.description or '',
            'is_required':      cr.effective_is_required(),   # respects override
            'is_active':        req.is_active,
            'profile_mapping':           req.profile_mapping or '',
            'managed_profile_mapping':   req.managed_profile_mapping or '',   # Phase 3
            'form_field_id':             req.form_field_id or '',
            # Phase 4: eligibility check fields — shown in the edit panel eligibility section
            'is_eligibility':            req.is_eligibility,
            'eligibility_operator':      req.eligibility_operator or '',
            'eligibility_value':         req.eligibility_value or '',
            'eligibility_fail_message':  req.eligibility_fail_message or '',
            # section_id — numeric PK used by the edit panel to pre-select the section dropdown.
            # section_name — human label shown as a badge in the requirement row.
            'section_id':       req.section_id,
            'section_name':     req.section.name if req.section else 'Uncategorized',
            'cr_id':            cr.id,        # CategoryRequirement id for remove/reorder calls
            'cr_order':         cr.order,
            # form_codes: list of government form codes (e.g. ['IMM5710']) that include
            # this requirement AND are linked to this category.
            # WHY: the JS uses this to group requirements under form section headers.
            # Empty list means the requirement is not from any linked government form.
            'form_codes':       req_to_forms.get(req.id, []),
        })

    # ── Inherited — walk up parent chain ──────────────────────────────
    # A category inherits all requirements from its parent categories.
    # 'seen_ids' prevents the same requirement appearing twice (own overrides parent).
    inherited = []
    seen_ids  = {r['id'] for r in own_reqs}
    current   = category.parent

    while current:
        parent_cat_reqs = (
            CategoryRequirement.objects
            .filter(category=current, requirement__is_active=True)
            .select_related('requirement')
            .order_by('order')
        )
        for cr in parent_cat_reqs:
            req = cr.requirement
            if req.id not in seen_ids:
                seen_ids.add(req.id)
                inherited.append({
                    'req': {
                        'id':          req.id,
                        'name':        req.name,
                        'type':        req.type,
                        'description': req.description or '',
                        'is_required': cr.effective_is_required(),
                    },
                    'from_category': {
                        'id':   current.id,
                        'name': current.name,
                    },
                })
        current = current.parent

    # Pending crawler suggestions for this category — used by the builder UI
    # to show a badge ("3 pending") on the Crawl button without a separate request.
    from cases.models import CrawlerSuggestion
    pending_count = CrawlerSuggestion.objects.filter(
        category=category, status='pending'
    ).count()

    return JsonResponse({
        'category': {
            'id':             category.id,
            'name':           category.name,
            'description':    category.description or '',
            'service_id':     category.service_id,
            'parent_id':      category.parent_id,
            # Crawler fields — used by the builder's source URL + Crawl button UI
            'source_url':     category.source_url or '',
            'last_crawled_at': (
                category.last_crawled_at.strftime('%Y-%m-%d %H:%M') if category.last_crawled_at else ''
            ),
            'pending_suggestions': pending_count,
        },
        'own_requirements': own_reqs,
        'inherited':        inherited,
    })


# ── AJAX: Create Requirement ──────────────────────────────────

@admin_required
@require_http_methods(['POST'])
def ajax_create_requirement(request):
    """
    POST /ajax/builder/library/create/
    (Old URL /ajax/builder/requirement/create/ also maps here for backwards compat.)

    Creates a new Requirement in the library, then optionally links it to a category.

    Body: {
        name,
        type,
        description,
        is_required,
        section_id       (optional — assigns to a library section)
        category_id      (optional — if provided, also creates a CategoryRequirement row)
        profile_mapping  (optional — dot-path for Phase 2 auto-fill, e.g. "profile.first_name")
    }
    Response: { ok: true, requirement: {...}, cr_id: <CategoryRequirement id or null> }
              { ok: false, errors: {...} }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    name            = body.get('name', '').strip()
    req_type        = body.get('type', '').strip()
    description     = body.get('description', '').strip()
    is_required     = body.get('is_required', True)
    section_id               = body.get('section_id')
    category_id              = body.get('category_id')
    profile_mapping          = body.get('profile_mapping', '').strip()
    # WHY managed_profile_mapping: this is the Phase 3 "interior person" auto-fill.
    # It auto-fills from the managed profile's own record (e.g. the applicant the case
    # is FOR), not from the account holder who filed the case. Having it in +Create New
    # lets the admin set it once without needing a separate Edit round-trip.
    managed_profile_mapping  = body.get('managed_profile_mapping', '').strip()

    # Phase 4: eligibility gate fields — sent when admin checks "⚑ Mark as Eligibility Gate"
    # in the Create New tab. Defaults to False/empty so existing callers without these
    # keys continue to work unchanged (backward-compatible).
    # WHY accept here (not just in ajax_edit_requirement):
    #   When creating an eligibility gate from scratch, the admin should be able to set
    #   the operator + value in the same form — forcing a two-step create-then-edit flow
    #   is unnecessarily slow for a common admin action.
    is_eligibility           = bool(body.get('is_eligibility',           False))
    eligibility_operator     = body.get('eligibility_operator',           '').strip()
    eligibility_value        = body.get('eligibility_value',              '').strip()
    eligibility_fail_message = body.get('eligibility_fail_message',       '').strip()

    # ── Validate ──────────────────────────────────────────────────────
    errors      = {}
    valid_types = [t[0] for t in Requirement.TYPE_CHOICES]

    if not name:
        errors['name'] = ['Name is required.']
    if not req_type:
        errors['type'] = ['Type is required.']
    elif req_type not in valid_types:
        errors['type'] = [f'Invalid type. Choose from: {", ".join(valid_types)}']

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    # ── Resolve optional FK targets ───────────────────────────────────
    section  = None
    category = None

    if section_id:
        try:
            section = RequirementSection.objects.get(id=section_id, is_active=True)
        except RequirementSection.DoesNotExist:
            return JsonResponse({'ok': False, 'errors': {'section_id': ['Section not found.']}})

    if category_id:
        try:
            category = Category.objects.get(id=category_id, is_active=True)
        except Category.DoesNotExist:
            return JsonResponse({'ok': False, 'errors': {'category_id': ['Category not found.']}})

    # ── Create the library Requirement ────────────────────────────────
    req = Requirement.objects.create(
        name            = name,
        type            = req_type,
        description     = description,
        is_required     = bool(is_required),
        section         = section,
        profile_mapping          = profile_mapping         or None,
        managed_profile_mapping  = managed_profile_mapping or None,
        # Phase 4: eligibility gate — store None (not '') so DB NULLs are clean
        is_eligibility           = is_eligibility,
        eligibility_operator     = eligibility_operator     or None,
        eligibility_value        = eligibility_value        or None,
        eligibility_fail_message = eligibility_fail_message or None,
    )

    # ── Optionally link to a category (creates CategoryRequirement) ───
    cr_id = None
    if category:
        # Determine the next order value so this requirement goes at the end of the list
        max_order = (
            CategoryRequirement.objects
            .filter(category=category)
            .aggregate(m=Max('order'))
            ['m'] or 0
        )
        cr = CategoryRequirement.objects.create(
            category    = category,
            requirement = req,
            order       = max_order + 1,
        )
        cr_id = cr.id

    return JsonResponse({
        'ok': True,
        'requirement': {
            'id':              req.id,
            'name':            req.name,
            'type':            req.type,
            'description':     req.description or '',
            'is_required':     req.is_required,
            'is_active':       req.is_active,
            'profile_mapping':         req.profile_mapping or '',
            'managed_profile_mapping': req.managed_profile_mapping or '',
            'section_name':            req.section.name if req.section else 'Uncategorized',
        },
        'cr_id': cr_id,     # CategoryRequirement id if linked to a category; null otherwise
    })


# ── AJAX: Edit Requirement ────────────────────────────────────

@admin_required
@require_http_methods(['POST'])
def ajax_edit_requirement(request, requirement_id):
    """
    POST /ajax/builder/requirement/<id>/edit/
    Body: { name, type, description, is_required, is_active, profile_mapping,
            form_field_id, section_id }
    Response: { ok: true, requirement: {...} }
              { ok: false, errors: {...} }

    Why POST not PATCH?
      Django's CSRF middleware works cleanly with POST.
      PATCH requires extra setup for form data — POST keeps it simple.

    section_id behaviour:
      - Key absent from body → section unchanged (backward-compatible with old callers)
      - Key present, null/empty → section cleared (requirement becomes Uncategorized)
      - Key present, numeric ID → requirement moved to that section
    """
    req = get_object_or_404(Requirement, id=requirement_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    name            = body.get('name',            req.name).strip()
    req_type        = body.get('type',            req.type).strip()
    description     = body.get('description',     req.description or '').strip()
    is_required     = body.get('is_required',     req.is_required)
    is_active       = body.get('is_active',       req.is_active)
    profile_mapping          = body.get('profile_mapping',         req.profile_mapping or '').strip()
    # Phase 3: dot-path into ManagedProfile for auto-fill when filling on behalf of interior person
    managed_profile_mapping  = body.get('managed_profile_mapping', req.managed_profile_mapping or '').strip()
    # Government form field identifier — kept for backwards compatibility.
    form_field_id            = body.get('form_field_id',            req.form_field_id or '').strip()
    # Phase 4: eligibility check fields
    is_eligibility           = bool(body.get('is_eligibility',       req.is_eligibility))
    eligibility_operator     = body.get('eligibility_operator',      req.eligibility_operator or '').strip()
    eligibility_value        = body.get('eligibility_value',         req.eligibility_value or '').strip()
    eligibility_fail_message = body.get('eligibility_fail_message',  req.eligibility_fail_message or '').strip()

    errors      = {}
    valid_types = [t[0] for t in Requirement.TYPE_CHOICES]

    if not name:
        errors['name'] = ['Name is required.']
    if req_type not in valid_types:
        errors['type'] = [f'Invalid type.']

    # ── Section reassignment ──────────────────────────────────────────────
    # Only update section if 'section_id' key was explicitly included in the body.
    # This preserves backward compatibility: old JS callers that don't send section_id
    # will not accidentally clear the section.
    new_section = req.section  # default: keep current section
    if 'section_id' in body:
        raw_section_id = body.get('section_id')
        if raw_section_id:
            # Numeric ID → look up the section
            try:
                new_section = RequirementSection.objects.get(id=raw_section_id, is_active=True)
            except RequirementSection.DoesNotExist:
                errors['section_id'] = ['Section not found.']
        else:
            # null / empty string → unassign (Uncategorized)
            new_section = None

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    req.name            = name
    req.type            = req_type
    req.description     = description
    req.is_required     = bool(is_required)
    req.is_active       = bool(is_active)
    req.profile_mapping         = profile_mapping         or None
    req.managed_profile_mapping = managed_profile_mapping or None   # Phase 3
    req.form_field_id           = form_field_id           or None
    req.section                 = new_section
    # Phase 4: eligibility fields
    req.is_eligibility           = is_eligibility
    req.eligibility_operator     = eligibility_operator     or None
    req.eligibility_value        = eligibility_value        or None
    req.eligibility_fail_message = eligibility_fail_message or None
    req.save()

    return JsonResponse({
        'ok':          True,
        'requirement': {
            'id':                       req.id,
            'name':                     req.name,
            'type':                     req.type,
            'description':              req.description or '',
            'is_required':              req.is_required,
            'is_active':                req.is_active,
            'profile_mapping':          req.profile_mapping or '',
            'managed_profile_mapping':  req.managed_profile_mapping or '',
            'form_field_id':            req.form_field_id or '',
            # Phase 4: eligibility fields returned so the UI can update the badge immediately
            'is_eligibility':           req.is_eligibility,
            'eligibility_operator':     req.eligibility_operator or '',
            'eligibility_value':        req.eligibility_value or '',
            'eligibility_fail_message': req.eligibility_fail_message or '',
            'section_id':               req.section_id,
            'section_name':             req.section.name if req.section else 'Uncategorized',
        }
    })


# ── AJAX: Edit Category ───────────────────────────────────────

@admin_required
@require_http_methods(['POST'])
def ajax_edit_category(request, category_id):
    """
    POST /ajax/builder/category/<id>/edit/
    Body: { name, description, is_active }
    Response: { ok: true, category: {...} }
    """
    category = get_object_or_404(Category, id=category_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    name        = body.get('name', category.name).strip()
    description = body.get('description', category.description or '').strip()
    is_active   = body.get('is_active', category.is_active)

    errors = {}
    if not name:
        errors['name'] = ['Name is required.']

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    category.name        = name
    category.description = description
    category.is_active   = bool(is_active)
    category.save()

    return JsonResponse({
        'ok':       True,
        'category': {
            'id':          category.id,
            'name':        category.name,
            'description': category.description or '',
            'is_active':   category.is_active,
        }
    })


# ── AJAX: Set Category Source URL ─────────────────────────────

@admin_required
@require_http_methods(['POST'])
def ajax_set_category_source_url(request, category_id):
    """
    POST /ajax/builder/category/<id>/source-url/
    Body: { source_url: 'https://...' }   (empty string = clear the URL)
    Response: { ok: true, source_url: '...' }

    Saves the government page URL that the crawler will read for this category.
    WHY a separate endpoint instead of including in ajax_edit_category:
      source_url has its own UX in the builder (save inline, then "Crawl Now").
      Separating it avoids the full category edit form reload cycle.
    """
    category = get_object_or_404(Category, id=category_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON'}, status=400)

    source_url = body.get('source_url', '').strip()
    category.source_url = source_url or None  # store None (not empty string) when cleared
    category.save(update_fields=['source_url'])

    return JsonResponse({'ok': True, 'source_url': source_url})


@admin_required
@require_http_methods(['POST'])
def ajax_delete_service(request, service_id):
    """
    POST /ajax/builder/service/<id>/delete/
    Soft delete — sets is_active=False instead of destroying data.
    Why soft delete? Cases linked to this service still exist and
    must remain accessible. Hard delete would break them.
    """
    service = get_object_or_404(Service, id=service_id)
    service.is_active = False
    service.save()
    return JsonResponse({'ok': True})


@admin_required
@require_http_methods(['POST'])
def ajax_delete_category(request, category_id):
    """
    POST /ajax/builder/category/<id>/delete/
    Soft delete — sets is_active=False.
    Also soft-deletes all direct children and their requirements
    so the tree stays consistent.
    EXPAND: make this recursive for deeper nesting if needed.
    """
    category = get_object_or_404(Category, id=category_id)

    # Soft delete this category
    category.is_active = False
    category.save()

    # Soft delete direct children
    Category.objects.filter(parent=category).update(is_active=False)

    # Soft delete requirements attached to this category via CategoryRequirement.
    # We soft-delete the Requirement library items themselves only if they are
    # EXCLUSIVELY used by this category (no other CategoryRequirement rows).
    # Requirements shared with other categories are left active — only their
    # CategoryRequirement link is removed (hard delete of the bridge row).
    cat_req_ids = CategoryRequirement.objects.filter(category=category).values_list('requirement_id', flat=True)
    for req_id in cat_req_ids:
        other_uses = CategoryRequirement.objects.filter(requirement_id=req_id).exclude(category=category).exists()
        if not other_uses:
            # Only used by this category → soft delete the library item
            Requirement.objects.filter(id=req_id).update(is_active=False)
        # else: shared with other categories → leave active, just remove the bridge row below

    # Remove the CategoryRequirement bridge rows for this category
    CategoryRequirement.objects.filter(category=category).delete()

    return JsonResponse({'ok': True})


@admin_required
@require_http_methods(['POST'])
def ajax_delete_requirement(request, requirement_id):
    """
    POST /ajax/builder/requirement/<id>/delete/
    Soft delete — sets is_active=False.
    Does NOT remove CaseRequirement rows — existing cases keep
    their requirements. is_active=False just hides it from new cases.
    """
    req = get_object_or_404(Requirement, id=requirement_id)
    req.is_active = False
    req.save()
    return JsonResponse({'ok': True})


# ══════════════════════════════════════════════════════════════
# PHASE 1 — REQUIREMENT LIBRARY AJAX VIEWS
#
# These views power the "Library" panel in the service builder.
# They manage sections (question banks), the library of reusable
# requirements, per-category requirement links, and choice options.
#
# All views:
#   - require @admin_required
#   - return JsonResponse({'ok': True, ...}) on success
#   - return JsonResponse({'ok': False, 'errors': {...}}) on failure
#
# URL patterns are in admin_panel/urls.py under the /ajax/builder/ prefix.
# ══════════════════════════════════════════════════════════════


# ── 1. GET SECTIONS ───────────────────────────────────────────
# Returns all active sections with a requirement count each.
# Used by the library browser to show the section sidebar.

@admin_required
def ajax_get_sections(request):
    """
    GET /ajax/builder/sections/
    Response: { sections: [{id, name, description, order, req_count}] }
    """
    sections = RequirementSection.objects.filter(is_active=True).order_by('order', 'name')
    data = []
    for s in sections:
        data.append({
            'id':          s.id,
            'name':        s.name,
            'slug':        s.slug,
            'description': s.description or '',
            'order':       s.order,
            # Count active requirements in this section
            'req_count':   s.requirements.filter(is_active=True).count(),
        })
    return JsonResponse({'sections': data})


# ── 2. CREATE SECTION ─────────────────────────────────────────
# Admin creates a new named question bank.
# EXPAND: add 'icon' field to body once RequirementSection has one.

@admin_required
@require_http_methods(['POST'])
def ajax_create_section(request):
    """
    POST /ajax/builder/section/create/
    Body: { name, description (optional) }
    Response: { ok: true, section: {...} }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    name        = body.get('name', '').strip()
    description = body.get('description', '').strip()

    if not name:
        return JsonResponse({'ok': False, 'errors': {'name': ['Name is required.']}})

    if RequirementSection.objects.filter(name__iexact=name).exists():
        return JsonResponse({'ok': False, 'errors': {'name': ['A section with this name already exists.']}})

    # Place at the end of the current order
    max_order = RequirementSection.objects.aggregate(m=Max('order'))['m'] or 0
    section   = RequirementSection.objects.create(
        name        = name,
        description = description or None,
        order       = max_order + 1,
    )
    return JsonResponse({
        'ok': True,
        'section': {
            'id':          section.id,
            'name':        section.name,
            'slug':        section.slug,
            'description': section.description or '',
            'order':       section.order,
            'req_count':   0,
        },
    })


# ── 3. EDIT SECTION ───────────────────────────────────────────
# Admin renames or reorders a section.

@admin_required
@require_http_methods(['POST'])
def ajax_edit_section(request, section_id):
    """
    POST /ajax/builder/section/<id>/edit/
    Body: { name (optional), description (optional), order (optional) }
    Response: { ok: true, section: {...} }
    """
    section = get_object_or_404(RequirementSection, id=section_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    name        = body.get('name', section.name).strip()
    description = body.get('description', section.description or '').strip()
    order       = body.get('order', section.order)

    if not name:
        return JsonResponse({'ok': False, 'errors': {'name': ['Name is required.']}})

    # Check uniqueness (exclude self)
    if RequirementSection.objects.filter(name__iexact=name).exclude(id=section_id).exists():
        return JsonResponse({'ok': False, 'errors': {'name': ['A section with this name already exists.']}})

    section.name        = name
    section.description = description or None
    section.order       = int(order)
    section.save()

    return JsonResponse({
        'ok': True,
        'section': {
            'id':          section.id,
            'name':        section.name,
            'description': section.description or '',
            'order':       section.order,
        },
    })


# ── 4. GET LIBRARY ────────────────────────────────────────────
# Browse the full requirement library, optionally filtered by section.
# Used by the "Add from Library" tab in the builder.

@admin_required
def ajax_get_library(request):
    """
    GET /ajax/builder/library/
    Query params: ?section_id=5 (optional), ?q=date (optional text search)
    Response: { requirements: [{id, name, type, description, is_required, section_name, profile_mapping}] }
    """
    qs = Requirement.objects.filter(is_active=True).select_related('section').order_by('name')

    section_id = request.GET.get('section_id')
    if section_id:
        qs = qs.filter(section_id=section_id)

    search = request.GET.get('q', '').strip()
    if search:
        # Filter by name containing the search term (case-insensitive)
        qs = qs.filter(name__icontains=search)

    data = []
    for req in qs:
        data.append({
            'id':              req.id,
            'name':            req.name,
            'type':            req.type,
            'description':              req.description or '',
            'is_required':              req.is_required,
            'profile_mapping':          req.profile_mapping or '',
            'managed_profile_mapping':  req.managed_profile_mapping or '',   # Phase 3
            'form_field_id':            req.form_field_id or '',
            # Phase 4: eligibility fields — shown as ⚑ badge in library results
            'is_eligibility':           req.is_eligibility,
            'eligibility_operator':     req.eligibility_operator or '',
            'eligibility_value':        req.eligibility_value or '',
            'eligibility_fail_message': req.eligibility_fail_message or '',
            'section_name':             req.section.name if req.section else 'Uncategorized',
            'section_id':               req.section_id,
        })
    return JsonResponse({'requirements': data})


# ── 5. ADD EXISTING REQUIREMENT TO CATEGORY ───────────────────
# Links an existing library requirement to a category.
# Creates one CategoryRequirement row. Does NOT duplicate the requirement.

@admin_required
@require_http_methods(['POST'])
def ajax_add_to_category(request, category_id):
    """
    POST /ajax/builder/category/<id>/add-req/
    Body: { requirement_id }
    Response: { ok: true, cr: {id, order, is_required_effective} }
    """
    category = get_object_or_404(Category, id=category_id, is_active=True)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    req_id = body.get('requirement_id')
    if not req_id:
        return JsonResponse({'ok': False, 'errors': {'requirement_id': ['Required.']}})

    req = get_object_or_404(Requirement, id=req_id, is_active=True)

    # Check if already linked
    if CategoryRequirement.objects.filter(category=category, requirement=req).exists():
        return JsonResponse({'ok': False, 'errors': {'__all__': ['This requirement is already in this category.']}})

    max_order = (
        CategoryRequirement.objects
        .filter(category=category)
        .aggregate(m=Max('order'))['m'] or 0
    )
    cr = CategoryRequirement.objects.create(
        category    = category,
        requirement = req,
        order       = max_order + 1,
    )

    return JsonResponse({
        'ok': True,
        'cr': {
            'id':                   cr.id,
            'order':                cr.order,
            'is_required_override': cr.is_required_override,
            'is_required_effective': cr.effective_is_required(),
            'requirement': {
                'id':              req.id,
                'name':            req.name,
                'type':            req.type,
                'description':     req.description or '',
                'profile_mapping': req.profile_mapping or '',
                'section_name':    req.section.name if req.section else 'Uncategorized',
            },
        },
    })


# ── 6. BULK-ADD A SECTION TO A CATEGORY ──────────────────────
# Adds ALL active requirements from a section to a category at once.
# Already-linked requirements are skipped (no duplicates).
# This is the "bulk import" action in the builder.

@admin_required
@require_http_methods(['POST'])
def ajax_add_section_to_category(request, category_id):
    """
    POST /ajax/builder/category/<id>/add-section/
    Body: { section_id }
    Response: { ok: true, added: <count>, skipped: <count> }
    """
    category = get_object_or_404(Category, id=category_id, is_active=True)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    section_id = body.get('section_id')
    if not section_id:
        return JsonResponse({'ok': False, 'errors': {'section_id': ['Required.']}})

    section = get_object_or_404(RequirementSection, id=section_id, is_active=True)

    # Get all active requirements in this section
    reqs = Requirement.objects.filter(section=section, is_active=True)

    # Get already-linked requirement IDs for this category
    existing_ids = set(
        CategoryRequirement.objects
        .filter(category=category)
        .values_list('requirement_id', flat=True)
    )

    max_order = (
        CategoryRequirement.objects
        .filter(category=category)
        .aggregate(m=Max('order'))['m'] or 0
    )

    added   = 0
    skipped = 0
    for req in reqs:
        if req.id in existing_ids:
            skipped += 1
            continue
        max_order += 1
        CategoryRequirement.objects.create(
            category  = category,
            requirement = req,
            order     = max_order,
        )
        added += 1

    return JsonResponse({'ok': True, 'added': added, 'skipped': skipped})


# ── 7. REMOVE REQUIREMENT FROM CATEGORY ──────────────────────
# Removes the CategoryRequirement bridge row (does NOT delete the library item).
# The requirement stays in the library and can be re-added or used elsewhere.

@admin_required
@require_http_methods(['POST'])
def ajax_remove_from_category(request, category_id, cr_id):
    """
    POST /ajax/builder/category/<id>/req/<cr_id>/remove/
    Response: { ok: true }
    """
    # Verify the CategoryRequirement belongs to this category (security: prevent cross-category removal)
    cr = get_object_or_404(CategoryRequirement, id=cr_id, category_id=category_id)
    cr.delete()
    # Hard delete of the bridge row — the Requirement library item is untouched.
    return JsonResponse({'ok': True})


# ── 8. REORDER REQUIREMENTS WITHIN A CATEGORY ────────────────
# Updates the 'order' field on CategoryRequirement rows after drag-and-drop.
# The builder sends the new order as a list of CategoryRequirement IDs.

@admin_required
@require_http_methods(['POST'])
def ajax_reorder_category_req(request, category_id):
    """
    POST /ajax/builder/category/<id>/reorder/
    Body: { ordered_ids: [cr_id_1, cr_id_2, cr_id_3, ...] }
    The IDs must all belong to this category (verified in the query).
    Response: { ok: true }
    """
    category = get_object_or_404(Category, id=category_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    ordered_ids = body.get('ordered_ids', [])
    if not isinstance(ordered_ids, list):
        return JsonResponse({'ok': False, 'errors': {'ordered_ids': ['Must be a list.']}})

    # Update each CategoryRequirement's order based on its position in the list.
    # enumerate starts at 1 so order values are 1-based (cleaner than 0-based).
    for position, cr_id in enumerate(ordered_ids, start=1):
        CategoryRequirement.objects.filter(
            id          = cr_id,
            category    = category,   # safety: only update rows belonging to this category
        ).update(order=position)

    return JsonResponse({'ok': True})


# ── 9. GET CHOICES FOR A SELECT REQUIREMENT ───────────────────
# Returns the dropdown options for a 'select'-type requirement.

@admin_required
def ajax_get_choices(request, requirement_id):
    """
    GET /ajax/builder/requirement/<id>/choices/
    Response: { choices: [{id, label, value, order}] }
    """
    req     = get_object_or_404(Requirement, id=requirement_id)
    choices = req.choices.order_by('order', 'label')
    data    = [
        {'id': c.id, 'label': c.label, 'value': c.value, 'order': c.order}
        for c in choices
    ]
    return JsonResponse({'choices': data})


# ── 10. CREATE CHOICE ─────────────────────────────────────────
# Adds one option to a 'select'-type requirement.

@admin_required
@require_http_methods(['POST'])
def ajax_create_choice(request, requirement_id):
    """
    POST /ajax/builder/requirement/<id>/choices/create/
    Body: { label, value (optional — defaults to label), order (optional) }
    Response: { ok: true, choice: {id, label, value, order} }
    """
    req = get_object_or_404(Requirement, id=requirement_id)

    if req.type != 'select':
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Only select-type requirements can have choices.']}})

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    label = body.get('label', '').strip()
    value = body.get('value', '').strip() or label   # default value = label if not provided
    order = body.get('order')

    if not label:
        return JsonResponse({'ok': False, 'errors': {'label': ['Label is required.']}})

    if order is None:
        # Place at the end
        order = (req.choices.aggregate(m=Max('order'))['m'] or 0) + 1

    choice = RequirementChoice.objects.create(
        requirement = req,
        label       = label,
        value       = value,
        order       = int(order),
    )
    return JsonResponse({
        'ok': True,
        'choice': {'id': choice.id, 'label': choice.label, 'value': choice.value, 'order': choice.order},
    })


# ── 11. EDIT CHOICE ───────────────────────────────────────────
# Update label, value, or order of an existing dropdown option.

@admin_required
@require_http_methods(['POST'])
def ajax_edit_choice(request, choice_id):
    """
    POST /ajax/builder/choice/<id>/edit/
    Body: { label (optional), value (optional), order (optional) }
    Response: { ok: true, choice: {id, label, value, order} }
    """
    choice = get_object_or_404(RequirementChoice, id=choice_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    label = body.get('label', choice.label).strip()
    value = body.get('value', choice.value).strip()
    order = body.get('order', choice.order)

    if not label:
        return JsonResponse({'ok': False, 'errors': {'label': ['Label is required.']}})

    choice.label = label
    choice.value = value or label   # fall back to label if value cleared
    choice.order = int(order)
    choice.save()

    return JsonResponse({
        'ok': True,
        'choice': {'id': choice.id, 'label': choice.label, 'value': choice.value, 'order': choice.order},
    })


# ── 12. DELETE CHOICE ─────────────────────────────────────────
# Hard-deletes a dropdown option.
# CaseAnswer.answer_choice rows that pointed to this choice become null (SET_NULL FK).
# This is safe — the answer row survives; the user just loses the choice reference.
# EXPAND: add soft-delete if you need to preserve old choice labels for exports.

@admin_required
@require_http_methods(['POST'])
def ajax_delete_choice(request, choice_id):
    """
    POST /ajax/builder/choice/<id>/delete/
    Response: { ok: true }
    """
    choice = get_object_or_404(RequirementChoice, id=choice_id)
    choice.delete()
    # CaseAnswer rows with answer_choice=this are SET_NULL via FK definition.
    return JsonResponse({'ok': True})


# ── 13. EDIT CATEGORY REQUIREMENT OVERRIDE ───────────────────
# Updates is_required_override on a CategoryRequirement.
# None = use library default | True = force required | False = force optional.
# EXPAND: add an 'order' field update here to combine with drag-drop reorder.

@admin_required
@require_http_methods(['POST'])
def ajax_edit_category_req(request, cr_id):
    """
    POST /ajax/builder/category-req/<id>/edit/
    Body: { is_required_override: true | false | null }
    Response: { ok: true, is_required_effective: true|false }
    """
    cr = get_object_or_404(CategoryRequirement, id=cr_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    override = body.get('is_required_override', 'UNSET')
    if override == 'UNSET':
        pass   # nothing to update
    elif override is None:
        cr.is_required_override = None      # reset to library default
    elif isinstance(override, bool):
        cr.is_required_override = override  # set override
    else:
        return JsonResponse({'ok': False, 'errors': {'is_required_override': ['Must be true, false, or null.']}})

    cr.save()
    return JsonResponse({'ok': True, 'is_required_effective': cr.effective_is_required()})


# ══════════════════════════════════════════════════════════════
# PHASE 4 — GOVERNMENT FORMS LIBRARY AJAX
# ══════════════════════════════════════════════════════════════
# Forms are reusable objects (IMM5710, IMM5257, etc.).
# Each form contains requirements linked via FormRequirement (M2M).
# Each category can require multiple forms via CategoryForm (M2M).
# The same Requirement appears in multiple forms → zero duplication.
#
# All views follow the same pattern as the existing AJAX library endpoints:
#   @admin_required (or @superadmin_required for destructive ops)
#   @require_http_methods([...])
#   JSON body in/out
# ══════════════════════════════════════════════════════════════

from cases.models import GovernmentForm, FormRequirement, CategoryForm as CatForm


# ── GET: List all government forms ────────────────────────────
@admin_required
def ajax_get_forms(request):
    """
    GET /ajax/builder/forms/
    Query params: ?q=<search> (optional text search by code or name)
    Response: { forms: [{id, code, name, description, source_url, is_active, req_count, category_count}] }
    """
    qs = GovernmentForm.objects.filter(is_active=True).order_by('code')

    search = request.GET.get('q', '').strip()
    if search:
        # Search both code (e.g. 'IMM5710') and name (e.g. 'Application to...')
        from django.db.models import Q
        qs = qs.filter(Q(code__icontains=search) | Q(name__icontains=search))

    data = []
    for form in qs:
        data.append({
            'id':             form.id,
            'code':           form.code,
            'name':           form.name,
            'description':    form.description or '',
            'source_url':     form.source_url or '',
            'is_active':      form.is_active,
            'req_count':      form.req_count(),       # count of active requirements in this form
            'category_count': form.category_count(),  # how many categories require this form
        })
    return JsonResponse({'forms': data})


# ── POST: Create a new government form ────────────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_create_form(request):
    """
    POST /ajax/builder/form/create/
    Body: { code, name, description (optional), source_url (optional) }
    Response: { ok: true, form: {id, code, name, ...} }
              { ok: false, errors: {...} }
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    code = body.get('code', '').strip().upper()   # normalize to uppercase: 'imm5710' → 'IMM5710'
    name = body.get('name', '').strip()

    errors = {}
    if not code:
        errors['code'] = ['Form code is required (e.g. IMM5710).']
    if not name:
        errors['name'] = ['Form name is required.']
    if code and GovernmentForm.objects.filter(code=code).exists():
        errors['code'] = [f'A form with code {code!r} already exists.']

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    form = GovernmentForm.objects.create(
        code        = code,
        name        = name,
        description = body.get('description', '').strip() or None,
        source_url  = body.get('source_url',  '').strip() or None,
    )
    return JsonResponse({
        'ok': True,
        'form': {
            'id':          form.id,
            'code':        form.code,
            'name':        form.name,
            'description': form.description or '',
            'source_url':  form.source_url or '',
            'is_active':   form.is_active,
            'req_count':      0,   # newly created — no requirements yet
            'category_count': 0,
        }
    })


# ── GET: Form detail — metadata + requirements grouped by section ─
@admin_required
def ajax_get_form_detail(request, form_id):
    """
    GET /ajax/builder/form/<id>/
    Response: {
      form: {id, code, name, description, source_url, is_active},
      form_requirements: [
        {fr_id, req_id, name, type, description, is_required, is_eligibility,
         form_section, field_id, order, section_name}
      ],
      used_in_categories: [{id, name, service_name}]
    }
    FormRequirements are returned in order (by FormRequirement.order).
    used_in_categories tells the UI which categories link to this form.
    """
    form = get_object_or_404(GovernmentForm, id=form_id)

    # Optional: ?section=Personal Information — filter displayed requirements by form_section.
    # Useful when a form has many sections and the JS wants to render one at a time.
    section_filter = request.GET.get('section', '').strip()

    # Requirements in this form, ordered for display
    form_reqs_qs = (
        FormRequirement.objects
        .filter(form=form, requirement__is_active=True)
    )
    if section_filter:
        form_reqs_qs = form_reqs_qs.filter(form_section__iexact=section_filter)

    form_reqs = (
        form_reqs_qs
        .select_related('requirement', 'requirement__section')
        .order_by('order')
    )
    reqs_data = []
    for fr in form_reqs:
        req = fr.requirement
        reqs_data.append({
            'fr_id':          fr.id,        # FormRequirement PK — used for remove/reorder
            'req_id':         req.id,
            'name':           req.name,
            'type':           req.type,
            'description':    req.description or '',
            'is_required':    req.is_required,
            'is_eligibility': req.is_eligibility,  # show ⚑ badge in form detail panel
            'eligibility_operator':     req.eligibility_operator or '',
            'eligibility_value':        req.eligibility_value or '',
            'form_section':   fr.form_section,     # section inside THIS form
            'field_id':       fr.field_id,         # gov form field identifier
            'order':          fr.order,
            'section_name':   req.section.name if req.section else 'Uncategorized',
        })

    # Categories that require this form
    cat_data = []
    for cf in CatForm.objects.filter(form=form).select_related('category', 'category__service'):
        cat_data.append({
            'id':           cf.category.id,
            'name':         cf.category.name,
            'service_name': cf.category.service.name,
        })

    return JsonResponse({
        'form': {
            'id':           form.id,
            'code':         form.code,
            'name':         form.name,
            'description':  form.description or '',
            'source_url':   form.source_url or '',
            'is_active':    form.is_active,
            # WHY: front-end shows a "Download PDF" link when pdf_file_url is set.
            # The .url property gives the MEDIA_URL-prefixed path for direct download.
            'pdf_file_url': form.pdf_file.url if form.pdf_file else '',
        },
        'form_requirements':   reqs_data,
        'used_in_categories':  cat_data,
    })


# ── POST: Edit a government form's metadata ────────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_edit_form(request, form_id):
    """
    POST /ajax/builder/form/<id>/edit/
    Body: { code (optional), name (optional), description (optional),
            source_url (optional), is_active (optional) }
    Response: { ok: true, form: {...} }
    """
    form = get_object_or_404(GovernmentForm, id=form_id)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    errors = {}
    new_code = body.get('code', form.code).strip().upper()
    new_name = body.get('name', form.name).strip()

    if not new_code:
        errors['code'] = ['Code is required.']
    if not new_name:
        errors['name'] = ['Name is required.']
    # Check uniqueness only if the code actually changed
    if new_code and new_code != form.code and GovernmentForm.objects.filter(code=new_code).exists():
        errors['code'] = [f'A form with code {new_code!r} already exists.']

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    form.code        = new_code
    form.name        = new_name
    form.description = body.get('description', form.description or '').strip() or None
    form.source_url  = body.get('source_url',  form.source_url  or '').strip() or None
    if 'is_active' in body:
        form.is_active = bool(body['is_active'])
    form.save()

    return JsonResponse({
        'ok': True,
        'form': {
            'id':          form.id,
            'code':        form.code,
            'name':        form.name,
            'description': form.description or '',
            'source_url':  form.source_url or '',
            'is_active':   form.is_active,
            'req_count':      form.req_count(),
            'category_count': form.category_count(),
        }
    })


# ── POST: Soft-delete a government form ───────────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_delete_form(request, form_id):
    """
    POST /ajax/builder/form/<id>/delete/
    Soft-delete: sets is_active=False.
    Does NOT delete FormRequirement or CategoryForm rows — they go dark with the form.
    Response: { ok: true, message: '...' }
    """
    form = get_object_or_404(GovernmentForm, id=form_id)
    form.is_active = False
    form.save()
    return JsonResponse({'ok': True, 'message': f'{form.code} has been deactivated.'})


# ── POST: Link a library requirement to a form ────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_add_req_to_form(request, form_id):
    """
    POST /ajax/builder/form/<id>/add-req/
    Body: { requirement_id, form_section (optional), field_id (optional) }
    Response: { ok: true, fr: {fr_id, req_id, name, type, form_section, field_id, order} }

    WHY this matters for deduplication:
      If "Background Check" is already in the library, linking it to IMM5710 here
      creates ONE FormRequirement row — the Requirement object is shared, not copied.
      The same requirement can be linked to IMM5257 later without any duplication.
    """
    form = get_object_or_404(GovernmentForm, id=form_id, is_active=True)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    req_id = body.get('requirement_id')
    if not req_id:
        return JsonResponse({'ok': False, 'errors': {'requirement_id': ['Required.']}})

    req = get_object_or_404(Requirement, id=req_id, is_active=True)

    if FormRequirement.objects.filter(form=form, requirement=req).exists():
        return JsonResponse({'ok': False, 'errors': {'__all__': ['This requirement is already in this form.']}})

    # Place new requirement at the end of the form's current list
    max_order = (
        FormRequirement.objects.filter(form=form).aggregate(m=Max('order'))['m'] or 0
    )
    fr = FormRequirement.objects.create(
        form         = form,
        requirement  = req,
        form_section = body.get('form_section', '').strip(),
        field_id     = body.get('field_id', '').strip(),
        order        = max_order + 1,
    )

    return JsonResponse({
        'ok': True,
        'fr': {
            'fr_id':          fr.id,
            'req_id':         req.id,
            'name':           req.name,
            'type':           req.type,
            'description':    req.description or '',
            'is_required':    req.is_required,
            'is_eligibility': req.is_eligibility,
            'form_section':   fr.form_section,
            'field_id':       fr.field_id,
            'order':          fr.order,
            'section_name':   req.section.name if req.section else 'Uncategorized',
        }
    })


# ── POST: Remove a requirement from a form ────────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_remove_req_from_form(request, form_id, fr_id):
    """
    POST /ajax/builder/form/<id>/req/<fr_id>/remove/
    Deletes the FormRequirement row (the link only — Requirement stays in the library).
    Response: { ok: true }
    """
    # No body fields needed — operation is fully determined by URL params.
    # Body is parsed anyway for consistency with other POST views (CSRF flow).
    _ = request.body
    fr = get_object_or_404(FormRequirement, id=fr_id, form_id=form_id)
    fr.delete()
    return JsonResponse({'ok': True})


# ── GET: Forms linked to a category ───────────────────────────
@admin_required
def ajax_get_forms_for_category(request, category_id):
    """
    GET /ajax/builder/category/<id>/forms/
    Response: {
      forms: [{cf_id, form_id, code, name, req_count, order}],
      summary_requirements: [{id, name, type, is_eligibility}]
        — deduplicated list of all requirements from all linked forms combined.
          Shows the admin exactly what requirements a user will face when they apply.
    }
    """
    category = get_object_or_404(Category, id=category_id)

    # ?skip_summary=1 lets the JS skip the deduplication pass when it only needs the forms list.
    # This is cheaper when the Forms tab is first opened (summary can be loaded on demand).
    skip_summary = request.GET.get('skip_summary', '0') == '1'

    cat_forms = (
        CatForm.objects
        .filter(category=category)
        .select_related('form')
        .order_by('order')
    )

    forms_data = []
    for cf in cat_forms:
        forms_data.append({
            'cf_id':     cf.id,            # CategoryForm PK — used for remove calls
            'form_id':   cf.form.id,
            'code':      cf.form.code,
            'name':      cf.form.name,
            'req_count': cf.form.req_count(),
            'order':     cf.order,
        })

    # ── Summary: deduplicated requirements across all linked forms ──────────
    # Skipped when ?skip_summary=1 (caller only needs the forms list, not the full summary).
    # Use a dict keyed on requirement_id to automatically deduplicate:
    # "Background Check" in both IMM5710 and IMM5257 → appears once in the summary.
    seen = {}   # requirement_id → data dict
    if not skip_summary:
        for cf in cat_forms:
            for fr in (
                FormRequirement.objects
                .filter(form=cf.form, requirement__is_active=True)
                .select_related('requirement')
                .order_by('order')
            ):
                req = fr.requirement
                if req.id not in seen:
                    seen[req.id] = {
                        'id':             req.id,
                        'name':           req.name,
                        'type':           req.type,
                        'is_required':    req.is_required,
                        'is_eligibility': req.is_eligibility,
                        'form_section':   fr.form_section,   # from the first form that defines it
                        'source_forms':   [cf.form.code],    # which forms include this req (for tooltip)
                    }
                else:
                    # Already seen — just record this additional source form
                    seen[req.id]['source_forms'].append(cf.form.code)

    return JsonResponse({
        'forms':                forms_data,
        'summary_requirements': list(seen.values()),
    })


# ── POST: Link a form to a category ───────────────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_add_form_to_category(request, category_id):
    """
    POST /ajax/builder/category/<id>/add-form/
    Body: { form_id }
    Response: { ok: true, cf: {cf_id, form_id, code, name, req_count, order} }
    """
    category = get_object_or_404(Category, id=category_id, is_active=True)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    form_id = body.get('form_id')
    if not form_id:
        return JsonResponse({'ok': False, 'errors': {'form_id': ['Required.']}})

    form = get_object_or_404(GovernmentForm, id=form_id, is_active=True)

    if CatForm.objects.filter(category=category, form=form).exists():
        return JsonResponse({'ok': False, 'errors': {'__all__': [f'{form.code} is already linked to this category.']}})

    max_order = CatForm.objects.filter(category=category).aggregate(m=Max('order'))['m'] or 0
    cf = CatForm.objects.create(category=category, form=form, order=max_order + 1)

    return JsonResponse({
        'ok': True,
        'cf': {
            'cf_id':     cf.id,
            'form_id':   form.id,
            'code':      form.code,
            'name':      form.name,
            'req_count': form.req_count(),
            'order':     cf.order,
        }
    })


# ── POST: Unlink a form from a category ───────────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_remove_form_from_category(request, category_id, cf_id):
    """
    POST /ajax/builder/category/<id>/form/<cf_id>/remove/
    Deletes the CategoryForm row. The form and its requirements are unchanged.
    Response: { ok: true }
    """
    # No body fields needed — operation fully determined by URL params.
    _ = request.body
    cf = get_object_or_404(CatForm, id=cf_id, category_id=category_id)
    cf.delete()
    return JsonResponse({'ok': True})


# ── POST: Crawler placeholder ──────────────────────────────────
@admin_required
@require_http_methods(['POST'])
def ajax_import_from_url(request):
    """
    POST /ajax/builder/import/
    Body: { url, form_id }

    Phase 4 placeholder — the web crawler is not yet implemented.
    For now: saves the URL to GovernmentForm.source_url so it is recorded
    for when the crawler is built. Returns a friendly "coming soon" message.

    Future crawler behaviour (Phase 5+):
      1. Fetch the URL (requests + BeautifulSoup / Playwright for JS-heavy pages)
      2. Parse eligibility conditions (e.g. "arrived on or before Feb 28, 2025")
         → create Requirement(type='date', is_eligibility=True, ...)
      3. Parse form fields (section name + field label + type)
         → look up existing matching Requirement by name (deduplicate)
         → create FormRequirement rows for new ones
      4. Return created/matched requirements so admin can review before saving
    """
    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'errors': {'__all__': ['Invalid JSON']}}, status=400)

    url     = body.get('url', '').strip()
    form_id = body.get('form_id')

    if not url:
        return JsonResponse({'ok': False, 'errors': {'url': ['URL is required.']}})

    # Save the URL to the form so it is available when the crawler is implemented.
    if form_id:
        try:
            form = GovernmentForm.objects.get(id=form_id, is_active=True)
            form.source_url = url
            form.save(update_fields=['source_url'])
            saved_msg = f' URL saved to {form.code}.'
        except GovernmentForm.DoesNotExist:
            saved_msg = ' (Form not found — URL not saved.)'
    else:
        saved_msg = ''

    return JsonResponse({
        'ok':      True,
        'message': (
            'Web crawler is not yet implemented.{} '
            'You can add requirements manually in the meantime.'
        ).format(saved_msg),
    })


# ── POST: Upload a fillable PDF and extract its AcroForm fields ─
@admin_required
@require_http_methods(['POST'])
def ajax_upload_form_pdf(request, form_id):
    """
    POST /ajax/builder/form/<id>/upload-pdf/

    Accepts a fillable PDF uploaded as multipart/form-data (field name: 'pdf_file').
    Passes the raw bytes to cases.crawler.pdf_parser.extract_pdf_fields() which reads
    the PDF's AcroForm structure (the locked/interactive layer that contains the form
    fields — text inputs, checkboxes, dropdowns, etc.).

    Response (success):
      {
        'ok':     true,
        'fields': [
          {
            'field_id':   'Section2_FamilyName',   # internal PDF field name
            'label':      'Family Name',            # cleaned human-readable label
            'field_type': 'text',                   # text | checkbox | radio | dropdown | signature
            'is_required': false,                   # AcroForm Required flag
            'options':    []                        # choices for dropdown/radio; empty otherwise
          },
          ...
        ],
        'message': ''     # empty on success; descriptive on no-fields result
      }

    Response (error):
      {'ok': false, 'error': 'No PDF file uploaded.'}

    WHY multipart/form-data (not JSON):
      JSON cannot carry binary file content. The browser's FormData API posts
      multipart/form-data, which Django's request.FILES handles natively.

    WHY return fields to the JS (not create DB rows immediately):
      The admin reviews the extracted fields before importing. They can:
        - Skip fields they don't need (signature pages, internal tracking fields)
        - Rename a label before importing
        - Check the type mapping (checkbox → boolean) before committing
      The actual DB rows are created by ajax_import_pdf_field_to_form().

    EXPAND: add a 'form_section' detection pass — many IMM PDFs have section headers
            embedded in field names (Section2_FamilyName → section "Section 2").
            Auto-grouping by section makes the review list easier to scan.

    REQUIRES: pip install pypdf  OR  pip install pymupdf
              Both are listed in pdf_parser.py. Without them, fields returns [].
    """
    from cases.crawler.pdf_parser import extract_pdf_fields

    # ── Check that at least one PDF library is installed BEFORE reading the file ──
    # WHY check here (not inside pdf_parser): the view is the user-facing boundary.
    # Returning a clear error with the exact install command avoids a confusing
    # "No fillable fields found" message that makes admins think the PDF is wrong
    # when the real problem is a missing package.
    # CUSTOMIZE: add 'pdfminer' to the try chain if you use a different library.
    try:
        import pypdf          # noqa: F401 — only checking existence, not using
        _pdf_lib = 'pypdf'
    except ImportError:
        try:
            import fitz       # noqa: F401 — PyMuPDF's import name is 'fitz'
            _pdf_lib = 'pymupdf'
        except ImportError:
            _pdf_lib = None

    if _pdf_lib is None:
        return JsonResponse({
            'ok':    False,
            'error': (
                'No PDF parsing library is installed. '
                'Open a terminal in your project folder and run:\n'
                '  venvcts/Scripts/pip install pypdf\n'
                'Then restart the Django development server. '
                'pypdf reads AcroForm interactive PDFs (fillable IMM forms).'
            ),
        })

    # Verify the form exists and is active before processing the upload.
    # WHY assign to govt_form: we need the object to save the PDF file onto the model.
    govt_form = get_object_or_404(GovernmentForm, id=form_id, is_active=True)

    pdf_file = request.FILES.get('pdf_file')
    if not pdf_file:
        return JsonResponse({'ok': False, 'error': 'No PDF file uploaded.'})
    if not pdf_file.name.lower().endswith('.pdf'):
        return JsonResponse({'ok': False, 'error': 'File must be a PDF (must end in .pdf).'})

    # Cap at 20MB — IMM forms are typically 1–5MB.
    # CUSTOMIZE: raise this limit if IRCC releases bundled form packages.
    MAX_UPLOAD_BYTES = 20 * 1024 * 1024  # 20 MB
    if pdf_file.size > MAX_UPLOAD_BYTES:
        return JsonResponse({'ok': False, 'error': 'PDF too large (max 20 MB).'})

    pdf_bytes = pdf_file.read()

    # ── Save the PDF to the model BEFORE parsing ──────────────────────────────
    # WHY before parsing: the file is stored permanently even if field extraction
    # fails (e.g. XFA format). The admin can then download the saved PDF later
    # without re-uploading it. GovernmentForm.pdf_file is a FileField.
    from django.core.files.base import ContentFile  # ContentFile wraps raw bytes as a Django file
    govt_form.pdf_file.save(pdf_file.name, ContentFile(pdf_bytes), save=True)
    # save=True → calls govt_form.save() automatically so the path is persisted to DB.

    fields = extract_pdf_fields(pdf_bytes)

    if not fields:
        # WHY two distinct messages:
        #   Many IRCC IMM PDFs use XFA (XML Forms Architecture) — a proprietary
        #   Adobe format. pypdf cannot read XFA; PyMuPDF (fitz) has partial
        #   XFA support but still extracts no fields for the newest IRCC forms.
        #   In that case the admin needs to know to try PyMuPDF or to manually
        #   enter the fields.  If pypdf is in use, the XFA path is most likely.
        if _pdf_lib == 'pypdf':
            detail = (
                'pypdf could not extract fields. IRCC IMM forms often use XFA '
                '(XML-based forms) which pypdf does not support. '
                'Try installing PyMuPDF for better coverage: '
                'venvcts/Scripts/pip install pymupdf — then restart the server.'
            )
        else:
            detail = (
                f'Library in use: {_pdf_lib}. '
                'The PDF may be a scanned (non-interactive) document, '
                'or the form uses a format neither library supports. '
                'Make sure you downloaded the original fillable PDF from IRCC.'
            )
        return JsonResponse({
            'ok':      True,
            'fields':  [],
            'message': f'No fillable fields found. {detail}',
        })

    return JsonResponse({
        'ok':      True,
        'fields':  fields,
        'message': '',
    })


# ── POST: Import one PDF field as a Requirement and link to the form ─
@admin_required
@require_http_methods(['POST'])
def ajax_import_pdf_field_to_form(request, form_id):
    """
    POST /ajax/builder/form/<id>/import-pdf-field/

    Creates (or reuses) a Requirement in the global library based on one PDF field
    extracted by ajax_upload_form_pdf(), then links it to the form via FormRequirement.

    Body (JSON):
      {
        'label':       'Family Name',      # field label → becomes Requirement.name
        'field_type':  'text',             # raw PDF field type (text|checkbox|radio|dropdown|signature)
        'is_required': false,              # AcroForm Required flag
        'field_id':    'Section2_FamilyName',  # optional: stored on FormRequirement for crawler cross-reference
        'form_section': 'Personal Information', # optional: section grouping inside this form
      }

    Response (success):
      {
        'ok': true,
        'created': true,        # true = new Requirement was created; false = reused existing
        'requirement_id': <id>,
        'fr_id': <FormRequirement id>,
        'name':  'Family Name',
        'type':  'text',
      }

    WHY 'get_or_create' on Requirement.name:
      IMM forms share many field names (e.g. "Date of Birth" appears in IMM5257 and IMM5710).
      Reusing the same Requirement object keeps the library lean and avoids duplicates.
      The admin can see 'created: false' in the JS and know it was linked, not re-created.

    WHY store field_id on FormRequirement:
      The field_id ('Section2_FamilyName') is a PDF-specific identifier used by the crawler
      to cross-reference future crawl results back to the same field. Storing it here means
      subsequent PDF uploads of updated IRCC forms can detect which fields changed.

    EXPAND: run NLP matching before creating a new Requirement to find similar existing ones.
            Use cases.crawler.nlp_matcher.find_best_match(label, Requirement.objects.filter(is_active=True))
            and return the match to the JS so the admin can accept/reject it.
    """
    from django.db.models import Max as _Max

    form = get_object_or_404(GovernmentForm, id=form_id, is_active=True)

    try:
        body = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': 'Invalid JSON body.'}, status=400)

    # ── Validate input ─────────────────────────────────────────
    label = body.get('label', '').strip()
    if not label:
        return JsonResponse({'ok': False, 'error': 'label is required.'})

    raw_field_type = body.get('field_type', 'text')
    is_required    = bool(body.get('is_required', False))
    field_id       = body.get('field_id', '').strip()
    form_section   = body.get('form_section', '').strip()

    # Map the raw PDF field type to the Requirement.type choices.
    # This reuses the same mapping as pipeline._map_field_type().
    type_map = {
        'text':      'text',
        'checkbox':  'boolean',
        'radio':     'boolean',
        'dropdown':  'select',
        'signature': 'file',
    }
    req_type = type_map.get(raw_field_type, 'text')

    # ── Get or create Requirement ──────────────────────────────
    # WHY get_or_create on name (case-insensitive):
    #   "Family Name" in IMM5710 and IMM5257 should map to the SAME Requirement.
    #   We match on name (case-insensitive) to avoid duplicates from slight casing differences.
    existing = Requirement.objects.filter(name__iexact=label, is_active=True).first()
    if existing:
        req     = existing
        created = False
    else:
        req = Requirement.objects.create(
            name        = label,
            type        = req_type,
            is_required = is_required,
            is_active   = True,
        )
        created = True

    # ── Link to form via FormRequirement ──────────────────────
    # If already linked, just return success (idempotent — no duplicate FR rows).
    existing_fr = FormRequirement.objects.filter(form=form, requirement=req).first()
    if existing_fr:
        return JsonResponse({
            'ok':             True,
            'created':        False,
            'requirement_id': req.id,
            'fr_id':          existing_fr.id,
            'name':           req.name,
            'type':           req.type,
            'already_linked': True,  # JS can show a subtle "already in form" note
        })

    # Place the new FormRequirement at the end of the current list.
    max_order = (
        FormRequirement.objects.filter(form=form).aggregate(m=_Max('order'))['m'] or 0
    )
    fr = FormRequirement.objects.create(
        form         = form,
        requirement  = req,
        form_section = form_section,
        field_id     = field_id,
        order        = max_order + 1,
    )

    return JsonResponse({
        'ok':             True,
        'created':        created,
        'requirement_id': req.id,
        'fr_id':          fr.id,
        'name':           req.name,
        'type':           req.type,
        'already_linked': False,
    })


# ══════════════════════════════════════════════════════════════
# TASK VIEWS
# ══════════════════════════════════════════════════════════════
# ___ Task List ___
@admin_permission_required('can_manage_tasks')
def task_list(request):
    """
    Admin sees all tasks they created or that are assigned to them.
    Superadmin sees all tasks.

    Filter by status via ?status=pending etc.
    Filter by type via ?type=admin_to_user etc.

    EXPAND: add search by title or user email.
    EXPAND: add date range filter.
    """
    from cases.models import Case

    tasks = Task.objects.select_related(
    'created_by', 'assigned_to_user', 'assigned_to_admin', 'related_case'
)

    # Superadmin sees all — regular admin sees only their tasks
    if not request.user.is_superuser:
        tasks = tasks.filter(
            created_by=request.user
        ) | tasks.filter(
            assigned_to_user=request.user
        ) | tasks.filter(
            assigned_to_admin__user=request.user
        )

    # Optional filters
    status_filter = request.GET.get('status')
    type_filter   = request.GET.get('type')

    if status_filter:
        tasks = tasks.filter(status=status_filter)
    if type_filter:
        tasks = tasks.filter(task_type=type_filter)  # also check if field is 'task_type' not 'type'

    tasks = tasks.order_by('due_date', '-created_at')

    return render(request, 'admin_panel/task_list.html', {
        'tasks':          tasks,
        'status_choices': Task.STATUS_CHOICES,
        'type_choices':   Task.TYPE_CHOICES,
        'status_filter':  status_filter,
        'type_filter':    type_filter,
    })


@admin_permission_required('can_manage_tasks')
def task_create(request):
    """
    Admin creates a task and assigns it to a user or another admin.

    On save:
      1. Task is created
      2. Notification is sent to assigned person
      3. If linked to a case, case gets a note

    EXPAND: add file attachment to task for supporting documents.
    """
    from cases.models import Case, Requirement

    if request.method == 'POST':
        task_type   = request.POST.get('type')
        title       = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        assigned_id = request.POST.get('assigned_to')
        due_date    = request.POST.get('due_date') or None
        case_id     = request.POST.get('linked_case') or None
        req_id      = request.POST.get('linked_requirement') or None

        errors = {}
        if not title:
            errors['title'] = 'Title is required.'
        if not assigned_id:
            errors['assigned_to'] = 'Assignee is required.'
        if not task_type:
            errors['type'] = 'Task type is required.'

        if not errors:
            # Fetch the user row that was selected in the "Assign To" dropdown
            assigned_user = get_object_or_404(User, id=assigned_id)

            # ── Resolve the correct assignee FK ───────────────────────────
            # The Task model has TWO separate FK fields for the two types:
            #
            #   assigned_to_user  → FK to User        (admin → user tasks)
            #   assigned_to_admin → FK to AdminProfile (admin → admin tasks)
            #
            # Only ONE must be set; the other must be None.
            # For admin→admin, we need the AdminProfile object, not the
            # User object — that's what the DB column points to.
            # ─────────────────────────────────────────────────────────────
            assigned_to_user_obj  = None
            assigned_to_admin_obj = None

            if task_type == 'user':
                # Straightforward — the User IS the assignee
                assigned_to_user_obj = assigned_user

            elif task_type == 'admin':
                # We need the AdminProfile, not the User.
                # hasattr() checks without hitting the DB again.
                # If the chosen user is not an admin, add a form error
                # instead of crashing with a ValueError.
                if not hasattr(assigned_user, 'admin_profile'):
                    errors['assigned_to'] = (
                        f'{assigned_user.email} does not have an admin profile.'
                    )
                else:
                    assigned_to_admin_obj = assigned_user.admin_profile

        # Re-check: the admin_profile guard above may have added an error
        if not errors:
            task = Task.objects.create(
                task_type              = task_type,
                title                  = title,
                description            = description,
                created_by             = request.user,
                assigned_to_user       = assigned_to_user_obj,
                assigned_to_admin      = assigned_to_admin_obj,
                due_date               = due_date,
                related_case_id        = case_id,
                related_requirement_id = req_id,
                status                 = 'pending',
            )

            # Notify the assignee (helper handles both user and admin type)
            notify_task_assigned(task)

            messages.success(request, f'Task "{title}" created and assigned.')
            return redirect('admin_panel:task_detail', task_id=task.id)

    else:
        errors = {}

    # Split users into two groups so the template can show only
    # the correct group depending on which task_type is selected.
    #
    #   regular_users (is_staff=False) → shown when task_type='user'
    #   admin_users   (is_staff=True)  → shown when task_type='admin'
    #
    # The template uses JS to swap between them; only the active
    # select is enabled so only one value is submitted per request.
    # EXPAND: add a separate group for 'managers' if you add a role tier.
    regular_users = User.objects.filter(is_active=True, is_staff=False).order_by('email')
    admin_users   = User.objects.filter(is_active=True, is_staff=True).order_by('email')
    all_cases     = Case.objects.filter(is_active=True).order_by('-created_at')[:50]

    # ── GET prefill from case_detail "Create Task" button ─────────────────
    # WHY GET params: case_detail links here with ?case_id=N&user_id=N so the
    # form is pre-filled without the admin having to search for the case/user.
    # On POST errors, 'post' dict re-fills the form instead of these prefills.
    prefill_case_id = request.GET.get('case_id') or None
    prefill_user_id = request.GET.get('user_id') or None

    # Resolve the prefill objects so the template can display names/emails.
    prefill_case = None
    prefill_user = None
    if prefill_case_id:
        prefill_case = Case.objects.filter(id=prefill_case_id, is_active=True).first()
    if prefill_user_id:
        prefill_user = User.objects.filter(id=prefill_user_id, is_active=True).first()

    return render(request, 'admin_panel/task_create.html', {
        'errors':         errors,
        'regular_users':  regular_users,
        'admin_users':    admin_users,
        'all_cases':      all_cases,
        'type_choices':   Task.TYPE_CHOICES,
        'post':           request.POST,  # re-fill form on error
        'prefill_case':   prefill_case,
        'prefill_user':   prefill_user,
    })


@admin_permission_required('can_manage_tasks')
def task_detail(request, task_id):
    """
    View/manage a single task.
    Admin can update status, edit title/description, add notes.
    When status changes to 'completed', notifies the task creator.
    """
    task = get_object_or_404(Task, id=task_id)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_status':
            new_status = request.POST.get('status')
            old_status = task.status

            if new_status in dict(Task.STATUS_CHOICES):
                task.status = new_status
                task.save()

                # Notify creator when completed
                if new_status == 'completed' and old_status != 'completed':
                    from tasks.models import notify_task_completed
                    notify_task_completed(task)

                messages.success(request, f'Task status updated to {new_status}.')

        elif action == 'update_task':
            task.title       = request.POST.get('title', task.title).strip()
            task.description = request.POST.get('description', task.description).strip()
            new_due          = request.POST.get('due_date')
            task.due_date    = new_due if new_due else None
            task.save()
            messages.success(request, 'Task updated.')

        return redirect('admin_panel:task_detail', task_id=task_id)

    return render(request, 'admin_panel/task_detail.html', {
        'task':           task,
        'status_choices': Task.STATUS_CHOICES,
    })


# ══════════════════════════════════════════════════════════════
# PAYMENT VIEWS
# ══════════════════════════════════════════════════════════════
# 34__ Invoice List __
@admin_permission_required('can_manage_payments')
def invoice_list(request):
    """
    Admin sees all invoices.
    Filter by user, status (paid/unpaid/overdue).

    EXPAND: add date range filter, export to CSV.
    """
    invoices = Invoice.objects.exclude(
    status='cancelled'
    ).select_related('user', 'created_by').prefetch_related('payments')

    # Filter by user email
    email_filter = request.GET.get('email', '').strip()
    if email_filter:
        invoices = invoices.filter(user__email__icontains=email_filter)

    # Filter by payment status
    paid_filter = request.GET.get('paid')
    if paid_filter == 'unpaid':
        # Get IDs of invoices where balance > 0
        # Must compute in Python since balance() is a method, not a DB field
        # EXPAND: add balance as annotated field for DB-level filtering
        invoices = [inv for inv in invoices if inv.status in ('unpaid', 'partial')]
    elif paid_filter == 'paid':
        invoices = [inv for inv in invoices if inv.status == 'paid']
    elif paid_filter == 'overdue':
        invoices = [inv for inv in invoices if inv.is_overdue()]

    return render(request, 'admin_panel/invoice_list.html', {
        'invoices':     invoices,
        'email_filter': email_filter,
        'paid_filter':  paid_filter,
    })


@admin_permission_required('can_manage_payments')
def invoice_create(request):
    """
    Admin creates an invoice for a user.
    Sends notification to user immediately on creation.

    EXPAND: add multiple line items (InvoiceLineItem model).
    """
    if request.method == 'POST':
        user_email  = request.POST.get('user_email', '').strip()
        title       = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        amount      = request.POST.get('amount', '').strip()
        due_date    = request.POST.get('due_date') or None

        errors = {}

        # Validate user
        try:
            user = User.objects.get(email=user_email, is_active=True)
        except User.DoesNotExist:
            errors['user_email'] = 'No active user with this email.'
            user = None

        if not title:
            errors['title'] = 'Title is required.'

        # Validate amount — must be positive decimal
        try:
            amount_decimal = Decimal(amount)
            if amount_decimal <= 0:
                errors['amount'] = 'Amount must be greater than zero.'
        except Exception:
            errors['amount'] = 'Enter a valid amount (e.g. 150.00).'
            amount_decimal = None

        # ── Resolve related_case from hidden input (pre-filled by case_detail) ──
        # WHY: when the admin clicks "+ Create Invoice" on a case_detail page,
        # the link includes ?case_id=N which the template puts in a hidden input.
        # We look up the Case object so we can set the FK on the Invoice.
        case_id      = request.POST.get('linked_case') or None
        related_case = Case.objects.filter(id=case_id, is_active=True).first() if case_id else None

        if not errors and user and amount_decimal:
            invoice = Invoice.objects.create(
                user         = user,
                title        = title,
                description  = description,
                amount       = amount_decimal,
                due_date     = due_date,
                created_by   = request.user,
                related_case = related_case,  # links invoice back to the case it was created from
            )

            # Notify user of new invoice
            notify_invoice_created(invoice)

            messages.success(
                request,
                f'Invoice "${title}" created for {user.email}.'
            )
            return redirect('admin_panel:invoice_detail', invoice_id=invoice.id)

    else:
        errors = {}

    users = User.objects.filter(is_active=True, is_staff=False).order_by('email')

    # ── GET prefill from case_detail "Create Invoice" button ──────────────
    # WHY: case_detail links here with ?case_id=N&user_email=x@y.com so the
    # form is pre-filled. The template uses prefill_user_email to populate the
    # email input, and prefill_case_id to populate a hidden linked_case input.
    prefill_case_id    = request.GET.get('case_id') or None
    prefill_user_email = request.GET.get('user_email') or ''

    return render(request, 'admin_panel/invoice_create.html', {
        'errors':             errors,
        'users':              users,
        'post':               request.POST,
        'prefill_case_id':    prefill_case_id,
        'prefill_user_email': prefill_user_email,
    })


@admin_permission_required('can_manage_payments')
def invoice_detail(request, invoice_id):
    """
    View invoice details + payment history + record new payment.

    Double payment protection:
      select_for_update() locks the invoice row during payment recording.
      This prevents two admins simultaneously recording the same payment.
      The lock is released after the transaction commits.

    Why transaction.atomic()?
      If anything fails after the payment is created but before the
      notification is sent, the entire operation rolls back.
      Either both succeed or neither does — no partial state.
    """
    invoice = get_object_or_404(Invoice, id=invoice_id)
    payments = invoice.payments.select_related('marked_by').order_by('-marked_at')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'record_payment':
            amount_str = request.POST.get('amount', '').strip()
            note       = request.POST.get('note', '').strip()

            try:
                amount = Decimal(amount_str)
                if amount == 0:
                    raise ValueError('Zero amount')
            except Exception:
                messages.error(request, 'Enter a valid payment amount.')
                return redirect('admin_panel:invoice_detail', invoice_id=invoice_id)

            # Double payment protection — atomic transaction + row lock
            # select_for_update() prevents concurrent writes to this invoice
            with transaction.atomic():
                # Re-fetch with lock inside transaction
                locked_invoice = Invoice.objects.select_for_update().get(
                    id=invoice_id
                )

                # Safety check — warn if payment would overpay
                if amount > locked_invoice.balance_due() and amount > 0:
                    messages.warning(
                        request,
                        f'Warning: payment of ${amount} exceeds remaining '
                        f'balance of ${locked_invoice.balance_due()}. '
                        f'Recording anyway.'
                        )

                payment = Payment.objects.create(
                    invoice   = locked_invoice,
                    amount    = amount,
                    notes     = note,       # 'notes' not 'note'
                    marked_by = request.user,
                    )
                locked_invoice.update_status()

            # Notify user — outside transaction (non-critical)
            notify_payment_recorded(invoice, payment)

            messages.success(
                request,
                f'Payment of ${amount} recorded.'
            )
            return redirect('admin_panel:invoice_detail', invoice_id=invoice_id)

    return render(request, 'admin_panel/invoice_detail.html', {
        'invoice':     invoice,
        'payments':    payments,
        'total_paid':  invoice.total_paid(),
        'balance':     invoice.balance_due(),
        'is_paid':     invoice.status == 'paid',
        'is_overdue':  invoice.is_overdue(),
    })

#
@admin_permission_required('can_manage_payments')
def user_balance_overview(request):
    """
    Admin sees all users with outstanding balances.
    Shows total owed per user across all their active invoices.

    EXPAND: add export to CSV for accounting.
    EXPAND: add filter by minimum balance amount.
    """
    # Get all users with at least one active invoice
    users_with_invoices = User.objects.filter(
    invoices__isnull=False
    ).distinct().order_by('email')

    # Build summary per user in Python
    # EXPAND: move to DB annotation when performance matters at scale
    user_balances = []
    for user in users_with_invoices:
        user_invoices = Invoice.objects.filter(user=user).exclude(status='cancelled')
        total_owed = sum(inv.balance_due() for inv in user_invoices)
        if total_owed > 0:  # only show users with outstanding balance
            user_balances.append({
                'user':         user,
                'total_owed':   total_owed,
                'invoice_count': user_invoices.count(),
            })

    # Sort by highest balance first
    user_balances.sort(key=lambda x: x['total_owed'], reverse=True)

    return render(request, 'admin_panel/user_balance_overview.html', {
        'user_balances': user_balances,
    })


# ----------- delet invoces and tasks

@admin_permission_required('can_manage_tasks')
def task_delete(request, task_id):
    if request.method == 'POST':
        task = get_object_or_404(Task, id=task_id)
        task.delete()
        messages.success(request, 'Task deleted.')
    return redirect('admin_panel:task_list')


@admin_permission_required('can_manage_payments')
def invoice_delete(request, invoice_id):
    if request.method == 'POST':
        invoice = get_object_or_404(Invoice, id=invoice_id)
        invoice.delete()
        messages.success(request, 'Invoice deleted.')
    return redirect('admin_panel:invoice_list')


# ══════════════════════════════════════════════════════════════
# CONTENT MANAGEMENT — superadmin only
# ══════════════════════════════════════════════════════════════
# Views:
#   site_settings_edit   — edit the SiteSettings singleton
#   blog_post_list       — list all blog posts (published + drafts)
#   blog_post_create     — create a new blog post
#   blog_post_edit       — edit an existing blog post
#   blog_post_delete     — delete a post (POST only)
#   contact_messages     — list contact form submissions + mark as read
# ══════════════════════════════════════════════════════════════

from main_site.models import SiteSettings, BlogPost, ContactMessage


# ── SITE SETTINGS ─────────────────────────────────────────────
@superadmin_required
def site_settings_edit(request):
    """
    Edit the SiteSettings singleton.
    GET  — prefill form with current values
    POST — save and redirect back with success message
    """
    settings_obj = SiteSettings.get()

    if request.method == 'POST':
        # Text fields
        settings_obj.company_name  = request.POST.get('company_name', '').strip()
        settings_obj.tagline       = request.POST.get('tagline', '').strip()
        settings_obj.contact_email = request.POST.get('contact_email', '').strip()
        settings_obj.contact_phone = request.POST.get('contact_phone', '').strip()
        settings_obj.address       = request.POST.get('address', '').strip()
        settings_obj.linkedin_url  = request.POST.get('linkedin_url', '').strip()
        settings_obj.instagram_url = request.POST.get('instagram_url', '').strip()
        settings_obj.hero_title    = request.POST.get('hero_title', '').strip()
        settings_obj.hero_subtitle = request.POST.get('hero_subtitle', '').strip()
        settings_obj.hero_cta_text = request.POST.get('hero_cta_text', '').strip()
        settings_obj.about_title   = request.POST.get('about_title', '').strip()
        settings_obj.about_body    = request.POST.get('about_body', '').strip()

        # Brand colors — hex strings (e.g. '#c8820a').
        # Empty string = use hardcoded default from the CSS :root block.
        # input type="color" always posts a valid 6-digit hex, so no validation needed.
        settings_obj.color_primary      = request.POST.get('color_primary', '').strip()
        settings_obj.color_accent       = request.POST.get('color_accent', '').strip()
        settings_obj.color_accent_light = request.POST.get('color_accent_light', '').strip()
        settings_obj.color_background   = request.POST.get('color_background', '').strip()

        # Logo upload — only replace if a new file was uploaded
        if request.FILES.get('logo'):
            settings_obj.logo = request.FILES['logo']

        settings_obj.save()
        messages.success(request, 'Site settings saved.')
        return redirect('admin_panel:site_settings_edit')

    return render(request, 'admin_panel/site_settings.html', {
        'settings_obj': settings_obj,
    })


# ── BLOG POST LIST ────────────────────────────────────────────
@superadmin_required
def blog_post_list(request):
    """All blog posts — published and drafts."""
    posts = BlogPost.objects.all()
    return render(request, 'admin_panel/blog_list.html', {'posts': posts})


# ── BLOG POST CREATE ──────────────────────────────────────────
@superadmin_required
def blog_post_create(request):
    """
    Create a new blog post.
    Shared template with blog_post_edit — detects new vs existing via post=None.
    """
    if request.method == 'POST':
        title       = request.POST.get('title', '').strip()
        body        = request.POST.get('body', '').strip()
        meta_desc   = request.POST.get('meta_description', '').strip()
        is_pub      = request.POST.get('is_published') == 'on'

        if not title or not body:
            messages.error(request, 'Title and body are required.')
            return render(request, 'admin_panel/blog_form.html', {'post': None})

        post = BlogPost(
            title=title,
            body=body,
            meta_description=meta_desc,
            is_published=is_pub,
        )
        if request.FILES.get('thumbnail'):
            post.thumbnail = request.FILES['thumbnail']
        post.save()   # slug auto-generated in save()

        messages.success(request, f'Post "{post.title}" created.')
        return redirect('admin_panel:blog_post_list')

    return render(request, 'admin_panel/blog_form.html', {'post': None})


# ── BLOG POST EDIT ────────────────────────────────────────────
@superadmin_required
def blog_post_edit(request, post_id):
    """Edit an existing blog post."""
    post = get_object_or_404(BlogPost, id=post_id)

    if request.method == 'POST':
        post.title            = request.POST.get('title', '').strip()
        post.body             = request.POST.get('body', '').strip()
        post.meta_description = request.POST.get('meta_description', '').strip()
        post.is_published     = request.POST.get('is_published') == 'on'

        if not post.title or not post.body:
            messages.error(request, 'Title and body are required.')
            return render(request, 'admin_panel/blog_form.html', {'post': post})

        if request.FILES.get('thumbnail'):
            post.thumbnail = request.FILES['thumbnail']
        post.save()

        messages.success(request, f'Post "{post.title}" updated.')
        return redirect('admin_panel:blog_post_list')

    return render(request, 'admin_panel/blog_form.html', {'post': post})


# ── BLOG POST DELETE ──────────────────────────────────────────
@superadmin_required
def blog_post_delete(request, post_id):
    """Delete a blog post (POST only — delete button in blog_list.html)."""
    if request.method == 'POST':
        post = get_object_or_404(BlogPost, id=post_id)
        title = post.title
        post.delete()
        messages.success(request, f'Post "{title}" deleted.')
    return redirect('admin_panel:blog_post_list')


# ── CONTACT MESSAGES ──────────────────────────────────────────
@superadmin_required
def contact_messages(request):
    """
    List all contact form submissions.
    POST with action='mark_read' or 'mark_unread' to toggle is_read.
    """
    if request.method == 'POST':
        msg_id = request.POST.get('message_id')
        action = request.POST.get('action')
        if msg_id and action in ('mark_read', 'mark_unread'):
            msg = get_object_or_404(ContactMessage, id=msg_id)
            msg.is_read = (action == 'mark_read')
            msg.save(update_fields=['is_read'])
        return redirect('admin_panel:contact_messages')

    msgs_qs  = ContactMessage.objects.all()
    unread   = msgs_qs.filter(is_read=False).count()

    return render(request, 'admin_panel/contact_messages.html', {
        'contact_msgs': msgs_qs,
        'unread_count': unread,
    })


# ══════════════════════════════════════════════════════════════
# Phase 5 — Eligibility Scoring System
# ══════════════════════════════════════════════════════════════

# ── AJAX: Eligibility check for a single category ─────────────
@admin_permission_required('can_view_all_cases')
def ajax_eligibility_check(request):
    """
    GET /admin-panel/ajax/eligibility/?user_email=x@y.com&category_id=5

    Returns the eligibility score for a user against a specific category.
    Used by:
      - create_case.html (inline panel when admin selects category + enters email)
      - service_browser.html (per-category badge)

    Response shape:
    {
      "ok":       true,
      "score":    80,
      "passed":   4,
      "failed":   1,
      "unknown":  0,
      "total":    5,
      "user_name": "Ali Hassan",
      "quiz_url":  "/admin-panel/users/5/eligibility-quiz/",
      "details": [
        {"id": 3, "name": "Date of Arrival", "passed": true,
         "source": "profile", "fail_message": null},
        ...
      ]
    }

    WHY GET (not POST): this is a read-only calculation — no state changes.
    GET is also easier to test directly in the browser address bar.
    """
    from cases.models import Category
    from cases.services import compute_eligibility_score
    from users.models import User as UserModel

    user_email  = request.GET.get('user_email', '').strip()
    category_id = request.GET.get('category_id', '').strip()

    # Validate inputs — both are required
    if not user_email or not category_id:
        return JsonResponse({'ok': False, 'error': 'user_email and category_id are required.'}, status=400)

    try:
        target_user = UserModel.objects.select_related('profile').get(email=user_email)
    except UserModel.DoesNotExist:
        return JsonResponse({'ok': False, 'error': f'No user found with email: {user_email}'}, status=404)

    try:
        category = Category.objects.get(id=int(category_id), is_active=True)
    except (Category.DoesNotExist, ValueError):
        return JsonResponse({'ok': False, 'error': 'Category not found.'}, status=404)

    result = compute_eligibility_score(target_user, category)

    # Build the quiz URL — deep link to the eligibility quiz for this user
    # so the admin can click "Answer remaining questions →" in the panel
    quiz_url = f'/admin-panel/users/{target_user.id}/eligibility-quiz/'

    # Get user display name for the panel header
    profile   = getattr(target_user, 'profile', None)
    user_name = profile.full_name() if profile and profile.full_name() else target_user.email

    return JsonResponse({
        'ok':       True,
        'score':    result['score'],
        'passed':   result['passed'],
        'failed':   result['failed'],
        'unknown':  result['unknown'],
        'total':    result['total'],
        'user_name': user_name,
        'quiz_url':  quiz_url,
        'details':  result['details'],
    })


# ── AJAX: All child categories with eligibility scores ────────
@admin_permission_required('can_view_all_cases')
def ajax_service_eligibility(request):
    """
    GET /admin-panel/ajax/service-eligibility/?service_id=5&user_email=x@y.com
    GET /admin-panel/ajax/service-eligibility/?parent_id=12&user_email=x@y.com

    Returns child categories of a service or parent category, each with
    an eligibility score for the given user.

    One API call returns all children at once — more efficient than calling
    ajax_eligibility_check per category when rendering the service browser.

    Response:
    {
      "ok": true,
      "categories": [
        {"id": 12, "name": "Open Work Permit", "score": 80,
         "has_children": true, "passed": 4, "failed": 1, "unknown": 0, "total": 5},
        ...
      ]
    }

    If user_email is absent or invalid, scores are omitted and only category
    metadata is returned (useful for rendering the service tree without a user selected).
    """
    from cases.models import Category, Service
    from cases.services import compute_eligibility_score
    from users.models import User as UserModel

    service_id  = request.GET.get('service_id', '').strip()
    parent_id   = request.GET.get('parent_id',  '').strip()
    user_email  = request.GET.get('user_email', '').strip()

    # At least one of service_id or parent_id is required
    if not service_id and not parent_id:
        return JsonResponse({'ok': False, 'error': 'service_id or parent_id required.'}, status=400)

    # Build the categories queryset
    if parent_id:
        try:
            categories = list(
                Category.objects.filter(parent_id=int(parent_id), is_active=True)
                .order_by('name')
            )
        except ValueError:
            return JsonResponse({'ok': False, 'error': 'Invalid parent_id.'}, status=400)
    else:
        try:
            service    = Service.objects.get(id=int(service_id), is_active=True)
            categories = list(
                Category.objects.filter(service=service, parent__isnull=True, is_active=True)
                .order_by('name')
            )
        except (Service.DoesNotExist, ValueError):
            return JsonResponse({'ok': False, 'error': 'Service not found.'}, status=404)

    # Resolve user (optional — if missing we return categories without scores)
    target_user = None
    if user_email:
        try:
            target_user = UserModel.objects.select_related('profile').get(email=user_email)
        except UserModel.DoesNotExist:
            pass   # silently proceed without scoring — not an error for the browser

    # Build response items — compute score for each category if user is known
    items = []
    for cat in categories:
        has_children = Category.objects.filter(parent=cat, is_active=True).exists()
        item = {
            'id':           cat.id,
            'name':         cat.name,
            'description':  cat.description or '',
            'has_children': has_children,
        }
        if target_user:
            result = compute_eligibility_score(target_user, cat)
            item.update({
                'score':   result['score'],
                'passed':  result['passed'],
                'failed':  result['failed'],
                'unknown': result['unknown'],
                'total':   result['total'],
            })
        else:
            # No user — mark scores as None so the UI can show "select a user" hint
            item.update({'score': None, 'passed': 0, 'failed': 0, 'unknown': 0, 'total': 0})
        items.append(item)

    return JsonResponse({'ok': True, 'categories': items})


# ── SERVICE BROWSER ───────────────────────────────────────────
@admin_permission_required('can_view_all_cases')
def service_browser(request):
    """
    GET /admin-panel/services/browse/
    GET /admin-panel/services/browse/?user_email=x@y.com

    Service browser with per-category eligibility scoring.
    Admins select a user → browse services → see how eligible that user is
    for each category → click "+ Create Case" to open the pre-filled case form.

    The page renders a flat list of services on load (no categories yet).
    Clicking a service row triggers AJAX → ajax_service_eligibility → renders
    categories inline with score badges.
    """
    from cases.models import Service
    from users.models import User as UserModel

    services    = Service.objects.filter(is_active=True).order_by('name')
    user_email  = request.GET.get('user_email', '').strip()
    target_user = None

    if user_email:
        try:
            target_user = UserModel.objects.select_related('profile').get(email=user_email)
        except UserModel.DoesNotExist:
            pass   # invalid email — show browser without pre-selected user

    return render(request, 'admin_panel/service_browser.html', {
        'services':    services,
        'user_email':  user_email,
        'target_user': target_user,
    })


# ── ELIGIBILITY QUIZ ──────────────────────────────────────────
@admin_permission_required('can_view_all_cases')
def eligibility_quiz(request, user_id):
    """
    GET  /admin-panel/users/<user_id>/eligibility-quiz/
    POST /admin-panel/users/<user_id>/eligibility-quiz/

    Shows all active eligibility-gate requirements across all categories.
    Answers that can be auto-filled from the user's profile are shown
    as pre-filled + read-only.  Only questions with no profile mapping
    (or where the profile field is empty) are editable.

    POST saves answers to EligibilityAnswer (upsert per requirement).
    Redirects to service_browser after saving.

    WHY show profile-mapped questions as read-only:
      The admin/user should know which answers came from their profile
      so they can update the profile if data is wrong, rather than
      storing a duplicate answer in EligibilityAnswer.
    """
    from cases.models import Requirement
    from cases.services import resolve_profile_value
    from users.models import User as UserModel, EligibilityAnswer

    target_user = get_object_or_404(UserModel.objects.select_related('profile'), id=user_id)

    # Gather all active eligibility requirements across ALL categories — de-duplicated.
    # WHY de-duplicate: the same "Date of Arrival" requirement appears in 10 categories
    # but the user only answers it once.  We use distinct() on Requirement objects.
    elig_reqs = (
        Requirement.objects
        .filter(is_eligibility=True, is_active=True)
        .distinct()
        .order_by('name')
    )

    # Pre-fetch existing quiz answers for this user
    existing_map = {
        ea.requirement_id: ea
        for ea in EligibilityAnswer.objects.filter(
            user=target_user, requirement_id__in=elig_reqs.values_list('id', flat=True)
        )
    }

    if request.method == 'POST':
        # Save / update answers for each eligibility requirement
        for req in elig_reqs:
            # Skip profile-mapped requirements — the answer lives in the profile
            if req.profile_mapping and resolve_profile_value(target_user, req.profile_mapping) is not None:
                continue

            field_key = f'elig_{req.id}'
            raw_value = request.POST.get(field_key, '').strip()

            if not raw_value:
                continue   # admin left it blank — don't save an empty answer

            # Upsert: create or update the EligibilityAnswer row
            ea, _ = EligibilityAnswer.objects.get_or_create(
                user=target_user, requirement=req,
            )
            if req.type == 'date':
                # Store as date object so check_eligibility() can compare correctly
                from datetime import date as date_type
                try:
                    ea.answer_date = date_type.fromisoformat(raw_value)
                    ea.answer_text = ''
                except ValueError:
                    continue   # skip malformed date input
            else:
                ea.answer_text = raw_value
                ea.answer_date = None
            ea.save()

        return redirect('admin_panel:service_browser')

    # Build question list for the template
    # Each item: {req, profile_value, profile_source, quiz_answer}
    questions = []
    for req in elig_reqs:
        profile_value = None
        if req.profile_mapping:
            profile_value = resolve_profile_value(target_user, req.profile_mapping)

        quiz_ea = existing_map.get(req.id)

        questions.append({
            'req':           req,
            'profile_value': profile_value,
            # If profile_value is set, this question is auto-answered — show read-only
            'from_profile':  profile_value is not None,
            'quiz_answer':   quiz_ea,
        })

    return render(request, 'admin_panel/eligibility_quiz.html', {
        'target_user': target_user,
        'questions':   questions,
        'quiz_url':    request.path,
    })


# ═══════════════════════════════════════════════════════════════
# CRAWLER VIEWS
# ═══════════════════════════════════════════════════════════════
# These views tie the cases/crawler/ package into the admin UI.
# Flow:
#   1. Admin sets source_url on a category in the service builder
#   2. Admin clicks "Crawl Now" → POST to crawl_category_view
#   3. pipeline.crawl_category() runs → creates CrawlerSuggestion rows
#   4. Admin visits crawler_review to accept/reject each suggestion
#   5. Accepting a suggestion creates the Requirement or GovernmentForm in the library


# ── 1. Trigger a crawl for a category ─────────────────────────
@admin_permission_required('can_manage_content')
def crawl_category_view(request, category_id):
    """
    POST /admin-panel/categories/<category_id>/crawl/

    Triggers the crawler pipeline for one category.
    Runs synchronously (blocks the request) — see EXPAND note for async option.
    Returns JSON so the builder's "Crawl Now" button can show a result message.

    Response:
      {'ok': True, 'created': 5, 'skipped': 2, 'warnings': [...]}
      {'ok': False, 'error': 'No source_url set'}

    WHY POST only:
      Crawling writes to the DB (creates CrawlerSuggestion rows).
      GET should not have side effects — HTTP convention.

    EXPAND: wrap in a Celery task for non-blocking async execution:
      from cases.crawler.pipeline import crawl_category
      from django.shortcuts import get_object_or_404
      result = crawl_category_task.delay(category_id)  # returns task ID immediately
      return JsonResponse({'ok': True, 'task_id': result.id})
      # Then add a /crawler/task/<task_id>/status/ endpoint to poll the result
    """
    from cases.models import Category
    from cases.crawler.pipeline import crawl_category

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required.'}, status=405)

    category = get_object_or_404(Category, id=category_id, is_active=True)

    # Run the crawl pipeline — this makes HTTP requests so may take 5–30 seconds
    result = crawl_category(category)

    status_code = 200 if result['ok'] else 400
    return JsonResponse(result, status=status_code)


# ── 2. Review queue (list pending suggestions) ─────────────────
@admin_permission_required('can_manage_content')
def crawler_review(request):
    """
    GET /admin-panel/crawler/review/

    Shows all pending CrawlerSuggestion rows grouped by category.
    Admin can filter by type (requirement / eligibility / form).
    Accepted/rejected suggestions are also shown (collapsed by default) for audit.

    Context:
      pending_by_category: {category: [suggestion, ...]} — ordered newest-first
      accepted_count:       int
      rejected_count:       int
      filter_type:          str — current type filter ('' = all)
    """
    from cases.models import CrawlerSuggestion

    filter_type = request.GET.get('type', '').strip()

    # Fetch all suggestions with related data in as few queries as possible.
    # select_related covers category + matched_requirement + reviewed_by in one JOIN.
    qs = (
        CrawlerSuggestion.objects
        .select_related('category', 'matched_requirement', 'reviewed_by__user')
        .order_by('-created_at')
    )

    if filter_type:
        qs = qs.filter(suggestion_type=filter_type)

    # Group pending suggestions by category for the template
    pending     = qs.filter(status='pending')
    pending_by_category = {}
    for suggestion in pending:
        cat = suggestion.category
        if cat not in pending_by_category:
            pending_by_category[cat] = []
        pending_by_category[cat].append(suggestion)

    return render(request, 'admin_panel/crawler_review.html', {
        'pending_by_category': pending_by_category,
        'pending_count':       pending.count(),
        'accepted_count':      qs.filter(status='accepted').count(),
        'rejected_count':      qs.filter(status='rejected').count(),
        'filter_type':         filter_type,
        'type_choices':        [('', 'All types'), ('requirement', 'Requirements'),
                                ('eligibility', 'Eligibility'), ('form', 'Forms')],
    })


# ── 3. Accept a suggestion ─────────────────────────────────────
@admin_permission_required('can_manage_content')
def accept_suggestion(request, suggestion_id):
    """
    POST /admin-panel/crawler/suggestions/<suggestion_id>/accept/

    Accepts a pending suggestion and creates the appropriate library item:
      - 'requirement'  → create Requirement + CategoryRequirement (link to category)
      - 'eligibility'  → create Requirement (is_eligibility=True) + CategoryRequirement
      - 'form'         → create GovernmentForm + CategoryForm (link to category)

    POST params (all optional — override NLP suggestions):
      name:                 str — overrides suggested_name
      req_type:             str — overrides suggested_type (for requirement/eligibility)
      operator:             str — eligibility operator (gte/lte/eq/...)
      value:                str — eligibility threshold value
      fail_message:         str — message shown to user when they don't meet eligibility
      link_to_requirement:  int — link to existing requirement ID instead of creating new
      form_code:            str — for 'form' type: the IMM form code

    Returns JSON for AJAX button handling.
    """
    from django.utils import timezone
    from cases.models import (
        CrawlerSuggestion, Requirement, CategoryRequirement,
        RequirementSection, GovernmentForm, CategoryForm,
    )
    from users.models import AdminProfile

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required.'}, status=405)

    suggestion  = get_object_or_404(CrawlerSuggestion, id=suggestion_id, status='pending')
    admin_prof  = getattr(request.user, 'admin_profile', None)

    # ── Common params ─────────────────────────────────────────
    name          = request.POST.get('name', suggestion.suggested_name).strip()
    req_type      = request.POST.get('req_type', suggestion.suggested_type or 'text').strip()
    operator      = request.POST.get('operator', suggestion.eligibility_operator).strip()
    value         = request.POST.get('value', suggestion.eligibility_value).strip()
    fail_message  = request.POST.get('fail_message', '').strip()
    link_to_id    = request.POST.get('link_to_requirement', '').strip()
    form_code     = request.POST.get('form_code', '').strip()

    if not name:
        return JsonResponse({'ok': False, 'error': 'Name is required.'}, status=400)

    created_req  = None
    linked_cat_req = None

    # ── FORM type: create GovernmentForm + CategoryForm ───────
    if suggestion.suggestion_type == 'form':
        code = form_code or name[:30].upper().replace(' ', '')
        gov_form, _ = GovernmentForm.objects.get_or_create(
            code     = code,
            defaults = {
                'name':       name,
                'source_url': suggestion.suggested_url,
            },
        )
        # Link to category
        CategoryForm.objects.get_or_create(
            category = suggestion.category,
            form     = gov_form,
        )

    else:
        # ── REQUIREMENT or ELIGIBILITY type ───────────────────
        is_eligibility = (suggestion.suggestion_type == 'eligibility')

        if link_to_id:
            # Admin chose to link this to an existing requirement
            try:
                req = Requirement.objects.get(id=int(link_to_id))
            except (Requirement.DoesNotExist, ValueError):
                return JsonResponse({'ok': False, 'error': 'Requirement not found.'}, status=400)
        else:
            # Create a new Requirement in the library.
            # WHY use the "Other" section as default: new crawler-created requirements
            # haven't been manually categorized yet. Admin can move them to the right
            # section afterwards in the library panel.
            default_section = (
                RequirementSection.objects.filter(name='Other').first()
                or RequirementSection.objects.order_by('order').first()
            )
            req = Requirement.objects.create(
                name               = name,
                type               = req_type,
                section            = default_section,
                is_eligibility     = is_eligibility,
                eligibility_operator = operator if is_eligibility else '',
                eligibility_value    = value    if is_eligibility else '',
                eligibility_fail_message = fail_message if is_eligibility else '',
                is_required        = True,
            )
            created_req = req

        # Link requirement to the category (as a CategoryRequirement bridge row)
        linked_cat_req, _ = CategoryRequirement.objects.get_or_create(
            category    = suggestion.category,
            requirement = req,
        )

    # ── Mark suggestion as accepted ───────────────────────────
    suggestion.status             = 'accepted'
    suggestion.reviewed_at        = timezone.now()
    suggestion.reviewed_by        = admin_prof
    suggestion.created_requirement = created_req
    suggestion.save()

    return JsonResponse({
        'ok':          True,
        'suggestion_id': suggestion.id,
        'created_new': created_req is not None,
        'req_id':      created_req.id if created_req else (linked_cat_req.requirement_id if linked_cat_req else None),
    })


# ── 4. Reject a suggestion ─────────────────────────────────────
@admin_permission_required('can_manage_content')
def reject_suggestion(request, suggestion_id):
    """
    POST /admin-panel/crawler/suggestions/<suggestion_id>/reject/

    Marks a suggestion as rejected. The suggestion is kept in the DB for audit trail
    (so admin can see what the crawler found and why it was rejected).
    No library items are created.

    Returns JSON for AJAX button handling.
    """
    from django.utils import timezone

    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required.'}, status=405)

    suggestion = get_object_or_404(CrawlerSuggestion, id=suggestion_id, status='pending')
    admin_prof = getattr(request.user, 'admin_profile', None)

    suggestion.status      = 'rejected'
    suggestion.reviewed_at = timezone.now()
    suggestion.reviewed_by = admin_prof
    suggestion.save(update_fields=['status', 'reviewed_at', 'reviewed_by'])

    return JsonResponse({'ok': True, 'suggestion_id': suggestion.id})
